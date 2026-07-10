#!/usr/bin/env python3
"""
硬件测试记录系统 - Flask 后端 API
"""
import os
import re
import io
import json
from datetime import datetime, date, timedelta
from functools import wraps
from zoneinfo import ZoneInfo
from urllib.parse import quote_plus

from flask import Flask, request, jsonify, send_file, send_from_directory, session
from flask_cors import CORS
from werkzeug.security import generate_password_hash, check_password_hash
from sqlalchemy import (
    create_engine, Column, Integer, String, Text, Date, DateTime, Enum as SAEnum,
    func, and_, or_, desc, asc, text, ForeignKey, UniqueConstraint, case
)
from sqlalchemy.orm import sessionmaker, declarative_base, relationship, joinedload
from sqlalchemy.exc import IntegrityError
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import markdown as md_lib

# ---------------------------------------------------------------------------
# 配置
# ---------------------------------------------------------------------------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR = os.path.dirname(BASE_DIR)
EXPORT_DIR = os.path.join(PROJECT_DIR, "exports")
os.makedirs(EXPORT_DIR, exist_ok=True)

# 加载 .env（本地开发 / 生产部署通用）
try:
    from dotenv import load_dotenv
    # 按优先级尝试多个路径：PROJECT_DIR > BASE_DIR > 当前工作目录
    env_candidates = [
        os.path.join(PROJECT_DIR, ".env"),
        os.path.join(BASE_DIR, ".env"),
        os.path.join(os.getcwd(), ".env"),
    ]
    loaded = False
    for p in env_candidates:
        if os.path.exists(p):
            load_dotenv(p)
            loaded = True
            break
    if not loaded:
        # 兜底：让 load_dotenv 自动搜索（从当前目录向上查找）
        load_dotenv()
except ImportError:
    pass

# 数据库配置 — 优先环境变量，回退默认值
try:
    _db_port = int(os.getenv("DB_PORT", "3306"))
except (ValueError, TypeError):
    print("[ERROR] DB_PORT 环境变量值无效，已回退到默认值 3306")
    _db_port = 3306

DB_CONFIG = {
    "host": os.getenv("DB_HOST", "127.0.0.1"),
    "port": _db_port,
    "user": os.getenv("DB_USER", "root"),
    "password": os.getenv("DB_PASSWORD", ""),
    "database": os.getenv("DB_DATABASE", "hardware_test_system"),
    "charset": "utf8mb4",
}

# Flask session 密钥 — 生产环境务必通过 .env/SECRET_KEY 设置
app = Flask(__name__, static_folder=None)
app.secret_key = os.getenv("SECRET_KEY", "dev-secret-change-in-production")
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"

DATABASE_URL = (
    f"mysql+pymysql://{DB_CONFIG['user']}:{quote_plus(DB_CONFIG['password'])}"
    f"@{DB_CONFIG['host']}:{DB_CONFIG['port']}/{DB_CONFIG['database']}"
    f"?charset={DB_CONFIG['charset']}"
)

# ---------------------------------------------------------------------------
# 时区工具 — 从系统配置读取，支持用户自行设置
# ---------------------------------------------------------------------------
_CN_TZ_CACHE = {"tz": None, "name": None}

def _get_system_tz():
    """获取系统配置的时区（带缓存），默认 Asia/Shanghai"""
    tz_name = "Asia/Shanghai"
    try:
        db = SessionLocal()
        try:
            cfg = db.query(SystemConfig).filter(SystemConfig.config_key == "timezone").first()
            if cfg and cfg.config_value and cfg.config_value.strip():
                tz_name = cfg.config_value.strip()
        finally:
            db.close()
    except Exception:
        pass
    # 缓存：仅在时区名变化时重新创建 ZoneInfo
    if _CN_TZ_CACHE["name"] != tz_name:
        try:
            _CN_TZ_CACHE["tz"] = ZoneInfo(tz_name)
            _CN_TZ_CACHE["name"] = tz_name
        except Exception:
            _CN_TZ_CACHE["tz"] = ZoneInfo("Asia/Shanghai")
            _CN_TZ_CACHE["name"] = "Asia/Shanghai"
    return _CN_TZ_CACHE["tz"]

def _now_cn():
    """返回当前系统时区时间（naive datetime）"""
    return datetime.now(_get_system_tz()).replace(tzinfo=None)

def _today_cn():
    """返回当前系统时区日期"""
    return datetime.now(_get_system_tz()).date()

# ---------------------------------------------------------------------------
# 工具函数
# ---------------------------------------------------------------------------
def parse_iso_date(date_str):
    """安全解析 ISO 日期字符串，统一转为系统时区（naive datetime）"""
    if not date_str:
        return None
    try:
        s = str(date_str).replace("Z", "+00:00")
        dt = datetime.fromisoformat(s)
        if dt.tzinfo is not None:
            dt = dt.astimezone(_get_system_tz()).replace(tzinfo=None)
        return dt
    except (ValueError, TypeError):
        try:
            return datetime.strptime(str(date_str)[:19], "%Y-%m-%dT%H:%M:%S")
        except (ValueError, TypeError):
            return None

# ---------------------------------------------------------------------------
# 数据库模型
# ---------------------------------------------------------------------------
engine = create_engine(DATABASE_URL, pool_size=10, pool_recycle=3600, pool_pre_ping=True, echo=False)
SessionLocal = sessionmaker(bind=engine)
Base = declarative_base()


class User(Base):
    __tablename__ = "users"

    id = Column(Integer, primary_key=True, autoincrement=True)
    username = Column(String(64), unique=True, nullable=False, comment="用户名")
    password_hash = Column(String(255), nullable=False, comment="密码哈希")
    display_name = Column(String(64), default="", comment="显示名称")
    role = Column(SAEnum("admin", "tester"), nullable=False, default="tester", comment="角色")
    is_active = Column(Integer, default=1, comment="是否启用")
    created_at = Column(DateTime, default=_now_cn)
    updated_at = Column(DateTime, default=_now_cn, onupdate=_now_cn)

    def to_dict(self):
        return {
            "id": self.id,
            "username": self.username,
            "display_name": self.display_name or self.username,
            "role": self.role,
            "is_active": self.is_active,
            "created_at": self.created_at.isoformat() if self.created_at else "",
            "updated_at": self.updated_at.isoformat() if self.updated_at else "",
        }


class Batch(Base):
    __tablename__ = "batches"

    id = Column(Integer, primary_key=True, autoincrement=True)
    batch_no = Column(String(32), unique=True, nullable=False, comment="批次号")
    description = Column(String(255), default="", comment="批次描述")
    operator = Column(String(64), nullable=False, comment="操作员")
    status = Column(SAEnum("planned", "in_progress", "paused", "completed", "cancelled"), default="planned", comment="批次状态")
    started_at = Column(DateTime, default=None, comment="测试开始时间")
    status_reason = Column(Text, default=None, comment="暂停/取消原因")
    device_count = Column(Integer, default=0, comment="批次内设备数量")
    total_count = Column(Integer, default=0, comment="板卡总数量")
    arrival_date = Column(DateTime, default=None, comment="到货时间")
    courier = Column(String(64), default="", comment="快递公司")
    tracking_no = Column(String(128), default="", comment="快递单号")
    test_date = Column(DateTime, default=None, comment="本次测试时间")
    test_standard = Column(String(255), default="", comment="测试标准")
    estimated_completion = Column(Date, default=None, comment="测试预计完成时间")
    actual_completion = Column(Date, default=None, comment="实际完成时间")
    completion_note = Column(Text, default="", comment="超期备注说明")
    created_at = Column(DateTime, default=_now_cn)
    updated_at = Column(DateTime, default=_now_cn, onupdate=_now_cn)

    devices = relationship("DeviceTest", back_populates="batch", lazy="dynamic")

    def to_dict(self):
        est = self.estimated_completion.isoformat() if self.estimated_completion else ""
        st = self.started_at.isoformat() if self.started_at else ""
        return {
            "id": self.id,
            "batch_no": self.batch_no,
            "description": self.description or "",
            "operator": self.operator,
            "status": self.status,
            "started_at": st,
            "status_reason": self.status_reason or "",
            "device_count": self.device_count,
            "total_count": self.total_count or 0,
            "arrival_date": self.arrival_date.isoformat() if self.arrival_date else "",
            "courier": self.courier or "",
            "tracking_no": self.tracking_no or "",
            "test_date": self.test_date.isoformat() if self.test_date else "",
            "test_standard": self.test_standard or "",
            "estimated_completion": est,
            "actual_completion": self.actual_completion.isoformat() if self.actual_completion else "",
            "completion_note": self.completion_note or "",
            "created_at": self.created_at.isoformat() if self.created_at else "",
            "updated_at": self.updated_at.isoformat() if self.updated_at else "",
            # Frontend alias fields
            "name": self.description or "",
            "standard_name": self.test_standard or "",
            "deadline": est,
            "start_date": st,
            "remark": self.completion_note or "",
        }


class DeviceTest(Base):
    __tablename__ = "device_tests"

    id = Column(Integer, primary_key=True, autoincrement=True)
    board_mac = Column(String(17), nullable=False, comment="板卡MAC地址")
    device_code = Column(String(64), default="", comment="设备编号（MAC去冒号大写）")
    device_type = Column(String(10), default="", comment="设备类型: 1=小周天 2=大周天 3=周天 4=智瞳")
    wireless_mac = Column(String(17), nullable=True, default=None, comment="无线MAC地址")
    ip_address = Column(String(45), nullable=True, default=None, comment="设备IP地址")
    status = Column(SAEnum("normal", "fault", "pending"), nullable=False, default="pending", comment="设备状态")
    fault_reason = Column(Text, comment="故障原因/现象")
    fault_disposition = Column(
        SAEnum("待返厂", "返厂中", "已返厂", "pending", "stored", ""), nullable=True, default="", comment="故障处置"
    )
    return_date = Column(Date, comment="返厂日期")
    return_courier = Column(String(64), default="", comment="返厂快递公司")
    return_tracking = Column(String(100), comment="返厂单号")
    fault_return_courier = Column(String(50), default="", comment="返厂跟踪-快递公司")
    fault_return_tracking = Column(String(50), default="", comment="返厂跟踪-返厂单编号")
    test_date = Column(DateTime, default=_now_cn, comment="测试时间")
    operator = Column(String(50), comment="测试操作员")
    notes = Column(Text, comment="备注")
    flow_status = Column(String(20), default="在库", comment="流转状态")
    flow_purpose = Column(String(100), default="", comment="流转用途")
    flow_destination = Column(String(200), default="", comment="流转目的地")
    flow_contact = Column(String(50), default="", comment="流转联系人")
    flow_phone = Column(String(30), default="", comment="联系电话")
    flow_out_date = Column(Date, nullable=True, comment="外出日期")
    flow_expected_return = Column(Date, nullable=True, comment="预计返还日期")
    flow_actual_return = Column(Date, nullable=True, comment="实际返还日期")
    flow_notes = Column(Text, comment="流转备注")
    sort_order = Column(Integer, default=None, comment="排序序号")
    batch_id = Column(Integer, ForeignKey("batches.id"), nullable=True, comment="所属批次")
    created_at = Column(DateTime, default=_now_cn)
    updated_at = Column(DateTime, default=_now_cn, onupdate=_now_cn)

    batch = relationship("Batch", back_populates="devices")

    __table_args__ = (
        UniqueConstraint("board_mac", "wireless_mac", "batch_id", name="uq_batch_board_wireless_mac"),
    )

    def to_dict(self):
        return {
            "id": self.id,
            "board_mac": self.board_mac,
            "device_code": self.device_code or "",
            "device_type": self.device_type or "",
            "wireless_mac": self.wireless_mac,
            "ip_address": self.ip_address,
            "status": self.status,
            "fault_reason": self.fault_reason or "",
            "fault_disposition": self.fault_disposition or "",
            "return_date": self.return_date.isoformat() if self.return_date else "",
            "return_courier": self.return_courier or "",
            "return_tracking": self.return_tracking or "",
            "fault_return_courier": self.fault_return_courier or "",
            "fault_return_tracking": self.fault_return_tracking or "",
            "test_date": self.test_date.isoformat() if self.test_date else "",
            "operator": self.operator or "",
            "notes": self.notes or "",
            "flow_status": self.flow_status or "在库",
            "flow_purpose": self.flow_purpose or "",
            "flow_destination": self.flow_destination or "",
            "flow_contact": self.flow_contact or "",
            "flow_phone": self.flow_phone or "",
            "flow_out_date": self.flow_out_date.isoformat() if self.flow_out_date else "",
            "flow_expected_return": self.flow_expected_return.isoformat() if self.flow_expected_return else "",
            "flow_actual_return": self.flow_actual_return.isoformat() if self.flow_actual_return else "",
            "flow_notes": self.flow_notes or "",
            "flow_overdue": bool(self.flow_expected_return and self.flow_expected_return < _today_cn() and self.flow_status and self.flow_status != "在库"),
            "sort_order": self.sort_order,
            "batch_id": self.batch_id,
            "batch_no": self.batch.batch_no if self.batch else "",
            "created_at": self.created_at.isoformat() if self.created_at else "",
            "updated_at": self.updated_at.isoformat() if self.updated_at else "",
        }


class TestStandard(Base):
    __tablename__ = "test_standards"

    id = Column(Integer, primary_key=True, autoincrement=True)
    name = Column(String(128), unique=True, nullable=False, comment="标准名称")
    description = Column(Text, default="", comment="标准描述")
    metrics = Column(Text, default="[]", comment="指标列表 JSON")
    created_at = Column(DateTime, default=_now_cn)
    updated_at = Column(DateTime, default=_now_cn, onupdate=_now_cn)

    def to_dict(self):
        return {
            "id": self.id,
            "name": self.name,
            "description": self.description or "",
            "metrics": json.loads(self.metrics) if self.metrics else [],
            "created_at": self.created_at.isoformat() if self.created_at else "",
            "updated_at": self.updated_at.isoformat() if self.updated_at else "",
        }


class DeviceTestResult(Base):
    __tablename__ = "device_test_results"

    id = Column(Integer, primary_key=True, autoincrement=True)
    device_id = Column(Integer, ForeignKey("device_tests.id"), nullable=False, comment="关联设备")
    standard_id = Column(Integer, ForeignKey("test_standards.id"), nullable=False, comment="关联测试标准")
    metric_name = Column(String(128), nullable=False, comment="指标名称")
    metric_value = Column(String(64), nullable=False, default="", comment="实测值")
    expected_value = Column(String(64), nullable=False, default="", comment="期望值")
    pass_ = Column("pass", Integer, nullable=False, default=0, comment="是否通过 0/1")
    notes = Column(Text, default=None, comment="备注")
    created_at = Column(DateTime, default=_now_cn)
    updated_at = Column(DateTime, default=_now_cn, onupdate=_now_cn)

    device = relationship("DeviceTest", backref="test_results")
    standard = relationship("TestStandard", backref="test_results")

    def to_dict(self):
        return {
            "id": self.id,
            "device_id": self.device_id,
            "standard_id": self.standard_id,
            "metric_name": self.metric_name,
            "metric_value": self.metric_value or "",
            "expected_value": self.expected_value or "",
            "pass": self.pass_ or 0,
            "notes": self.notes or "",
            "created_at": self.created_at.isoformat() if self.created_at else "",
            "updated_at": self.updated_at.isoformat() if self.updated_at else "",
        }


class OperatorHistory(Base):
    __tablename__ = "operators"

    id = Column(Integer, primary_key=True, autoincrement=True)
    name = Column(String(50), unique=True, nullable=False, comment="操作员名称")
    use_count = Column(Integer, default=0, comment="使用次数")
    last_used = Column(DateTime, default=_now_cn, comment="最近使用时间")

    def to_dict(self):
        return {
            "id": self.id,
            "name": self.name,
            "use_count": self.use_count,
            "last_used": self.last_used.isoformat() if self.last_used else "",
        }


class SystemConfig(Base):
    __tablename__ = "system_configs"

    id = Column(Integer, primary_key=True, autoincrement=True)
    config_key = Column(String(128), unique=True, nullable=False, comment="配置键")
    config_value = Column(Text, default="", comment="配置值")
    created_at = Column(DateTime, default=_now_cn)
    updated_at = Column(DateTime, default=_now_cn, onupdate=_now_cn)

    def to_dict(self):
        return {
            "id": self.id,
            "config_key": self.config_key,
            "config_value": self.config_value,
            "created_at": self.created_at.isoformat() if self.created_at else "",
            "updated_at": self.updated_at.isoformat() if self.updated_at else "",
        }


class ReturnRecord(Base):
    __tablename__ = "return_records"

    id = Column(Integer, primary_key=True, autoincrement=True)
    return_code = Column(String(50), default="", comment="返厂单编号")
    tracking_no = Column(String(50), unique=True, nullable=False, comment="快递单号（主业务键）")
    courier = Column(String(50), default="", comment="快递公司")
    status = Column(String(20), default="进行中", comment="返厂状态：进行中/完成/取消")
    device_count = Column(Integer, default=0, comment="关联设备数")
    created_at = Column(DateTime, default=_now_cn)
    updated_at = Column(DateTime, default=_now_cn, onupdate=_now_cn)

    devices = relationship("DeviceReturnRecord", back_populates="return_record", lazy="dynamic")

    def to_dict(self):
        return {
            "id": self.id,
            "return_code": self.return_code or "",
            "tracking_no": self.tracking_no,
            "courier": self.courier or "",
            "status": self.status,
            "device_count": self.device_count,
            "created_at": self.created_at.isoformat() if self.created_at else "",
            "updated_at": self.updated_at.isoformat() if self.updated_at else "",
        }

    def to_detail_dict(self):
        d = self.to_dict()
        d["devices"] = [dr.device.to_dict() if dr.device else None for dr in self.devices.all()]
        d["devices"] = [dev for dev in d["devices"] if dev is not None]
        return d


class DeviceReturnRecord(Base):
    __tablename__ = "device_return_records"

    id = Column(Integer, primary_key=True, autoincrement=True)
    device_id = Column(Integer, ForeignKey("device_tests.id"), nullable=False, comment="设备ID")
    return_record_id = Column(Integer, ForeignKey("return_records.id"), nullable=False, comment="返厂记录ID")
    __table_args__ = (
        UniqueConstraint("device_id", "return_record_id", name="uq_device_return"),
    )

    device = relationship("DeviceTest", backref="return_records_link")
    return_record = relationship("ReturnRecord", back_populates="devices")


class TestCommand(Base):
    __tablename__ = "test_commands"
    id = Column(Integer, primary_key=True, autoincrement=True)
    title = Column(String(128), nullable=False, comment="命令标题")
    command = Column(Text, nullable=False, comment="命令内容")
    category = Column(String(32), default="其他", comment="分类：网络测试/硬件检测/系统命令/诊断调试/其他")
    notes = Column(Text, default="", comment="说明/注意事项")
    sort_order = Column(Integer, default=0, comment="排序")
    created_at = Column(DateTime, default=_now_cn)
    updated_at = Column(DateTime, default=_now_cn, onupdate=_now_cn)

    def to_dict(self):
        return {
            "id": self.id,
            "title": self.title,
            "command": self.command,
            "category": self.category,
            "notes": self.notes or "",
            "sort_order": self.sort_order,
            "created_at": self.created_at.isoformat() if self.created_at else "",
            "updated_at": self.updated_at.isoformat() if self.updated_at else ""
        }


class UserCommandPin(Base):
    __tablename__ = "user_command_pins"
    id = Column(Integer, primary_key=True, autoincrement=True)
    user_id = Column(Integer, ForeignKey("users.id"), nullable=False, comment="用户ID")
    command_id = Column(Integer, ForeignKey("test_commands.id"), nullable=False, comment="命令ID")
    sort_order = Column(Integer, default=0, comment="置顶排序")
    created_at = Column(DateTime, default=_now_cn)

    __table_args__ = (
        UniqueConstraint('user_id', 'command_id', name='uq_user_command_pin'),
    )


# ---------------------------------------------------------------------------
# Flask 应用初始化
# ---------------------------------------------------------------------------
CORS(app, supports_credentials=True)


def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


def api_response(success=True, data=None, message="OK", code=200):
    resp = {"success": success, "message": message, "data": data}
    return jsonify(resp), code


# ---------------------------------------------------------------------------
# 认证装饰器
# ---------------------------------------------------------------------------
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session:
            return api_response(success=False, message="请先登录", code=401)
        return f(*args, **kwargs)
    return decorated


def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session:
            return api_response(success=False, message="请先登录", code=401)
        if session.get("role") != "admin":
            return api_response(success=False, message="需要管理员权限", code=403)
        return f(*args, **kwargs)
    return decorated


# ---------------------------------------------------------------------------
# 验证工具
# ---------------------------------------------------------------------------
MAC_PATTERN = re.compile(r"^([0-9A-Fa-f]{2}[:-]){5}([0-9A-Fa-f]{2})$")
IP_PATTERN = re.compile(
    r"^(25[0-5]|2[0-4]\d|1\d{2}|[1-9]?\d)"
    r"(\.(25[0-5]|2[0-4]\d|1\d{2}|[1-9]?\d)){3}$"
)


def normalize_mac(raw):
    """将紧凑十六进制MAC（如 'AAD37A769C3F'）规范化为冒号分隔格式。
    已是冒号/横杠分隔的格式原样返回。"""
    if not raw:
        return raw
    s = raw.strip().upper()
    if ':' in s or '-' in s:
        return s
    if len(s) == 12 and all(c in '0123456789ABCDEF' for c in s):
        return ':'.join(s[i:i+2] for i in range(0, 12, 2))
    return s


def validate_device_data(data, is_update=False):
    errors = []
    if not is_update:
        mac = normalize_mac(data.get("board_mac") or "")
        if not mac or not MAC_PATTERN.match(mac):
            errors.append("板卡MAC地址格式无效 (应为 XX:XX:XX:XX:XX:XX)")
        wmac = normalize_mac(data.get("wireless_mac") or "")
        if wmac and not MAC_PATTERN.match(wmac):
            errors.append("无线MAC地址格式无效 (应为 XX:XX:XX:XX:XX:XX)")
        ip = (data.get("ip_address") or "").strip()
        if ip and not IP_PATTERN.match(ip):
            errors.append("IP地址格式无效")
    else:
        if "board_mac" in data:
            mac = normalize_mac(data["board_mac"])
            if mac and not MAC_PATTERN.match(mac):
                errors.append("板卡MAC地址格式无效")
        if "wireless_mac" in data:
            wmac = normalize_mac(data["wireless_mac"])
            if wmac and not MAC_PATTERN.match(wmac):
                errors.append("无线MAC地址格式无效")
        if "ip_address" in data:
            ip = data["ip_address"].strip()
            if ip and not IP_PATTERN.match(ip):
                errors.append("IP地址格式无效")

    status = data.get("status", "")
    if status == "fault":
        if not (data.get("fault_reason") or "").strip():
            errors.append("故障状态时必须填写故障原因或现象")
        disposition = data.get("fault_disposition", "")
        # 待返厂只需标记处置类型，不需要强制填写返厂日期/快递/单号
        if disposition == "待返厂":
            pass
    return errors


# ===========================================================================
# 认证 API
# ===========================================================================

@app.route("/api/auth/login", methods=["POST"])
def auth_login():
    data = request.get_json(force=True) or {}
    username = data.get("username", "").strip()
    password = data.get("password", "")

    if not username or not password:
        return api_response(success=False, message="用户名和密码不能为空", code=400)

    db = next(get_db())
    try:
        user = db.query(User).filter(User.username == username).first()
        if not user or not check_password_hash(user.password_hash, password):
            return api_response(success=False, message="用户名或密码错误", code=401)
        if not user.is_active:
            return api_response(success=False, message="账号已被禁用", code=403)

        session["user_id"] = user.id
        session["username"] = user.username
        session["display_name"] = user.display_name or user.username
        session["role"] = user.role
        session.permanent = True

        return api_response(data={
            "id": user.id,
            "username": user.username,
            "display_name": user.display_name or user.username,
            "role": user.role,
        }, message="登录成功")
    finally:
        db.close()


@app.route("/api/auth/logout", methods=["POST"])
def auth_logout():
    session.clear()
    return api_response(message="已登出")


@app.route("/api/auth/me", methods=["GET"])
def auth_me():
    if "user_id" not in session:
        return api_response(success=False, message="未登录", code=401)
    return api_response(data={
        "id": session["user_id"],
        "username": session["username"],
        "display_name": session.get("display_name", session["username"]),
        "role": session.get("role"),
    })


# ===========================================================================
# 用户管理 API（仅 admin）
# ===========================================================================

@app.route("/api/users", methods=["GET"])
@admin_required
def list_users():
    db = next(get_db())
    try:
        users = db.query(User).order_by(User.id).all()
        return api_response(data=[u.to_dict() for u in users])
    finally:
        db.close()


@app.route("/api/users", methods=["POST"])
@admin_required
def create_user():
    data = request.get_json(force=True) or {}
    username = data.get("username", "").strip()
    password = data.get("password", "").strip()
    display_name = data.get("display_name", "").strip()
    role = data.get("role", "tester")

    if not username or not password:
        return api_response(success=False, message="用户名和密码不能为空", code=400)
    if role not in ("admin", "tester"):
        return api_response(success=False, message="无效的角色", code=400)

    db = next(get_db())
    try:
        if db.query(User).filter(User.username == username).first():
            return api_response(success=False, message="用户名已存在", code=409)

        user = User(
            username=username,
            password_hash=generate_password_hash(password, method='pbkdf2:sha256'),
            display_name=display_name or username,
            role=role,
        )
        db.add(user)
        db.commit()
        db.refresh(user)
        return api_response(data=user.to_dict(), message="用户创建成功", code=201)
    except Exception as e:
        db.rollback()
        return api_response(success=False, message=str(e), code=500)
    finally:
        db.close()


@app.route("/api/users/<int:user_id>", methods=["PUT"])
@admin_required
def update_user(user_id):
    data = request.get_json(force=True) or {}
    db = next(get_db())
    try:
        user = db.query(User).filter(User.id == user_id).first()
        if not user:
            return api_response(success=False, message="用户不存在", code=404)

        if "username" in data:
            new_username = data["username"].strip()
            if new_username and new_username != user.username:
                existing = db.query(User).filter(User.username == new_username).first()
                if existing:
                    return api_response(success=False, message="用户名已存在", code=409)
                user.username = new_username
        if "display_name" in data:
            user.display_name = data["display_name"].strip()
        if "password" in data and data["password"]:
            user.password_hash = generate_password_hash(data["password"], method='pbkdf2:sha256')
        if "role" in data and data["role"] in ("admin", "tester"):
            user.role = data["role"]
        if "is_active" in data:
            user.is_active = 1 if data["is_active"] else 0

        db.commit()
        db.refresh(user)
        return api_response(data=user.to_dict(), message="用户更新成功")
    except Exception as e:
        db.rollback()
        return api_response(success=False, message=str(e), code=500)
    finally:
        db.close()


@app.route("/api/users/<int:user_id>", methods=["DELETE"])
@admin_required
def delete_user(user_id):
    if user_id == session.get("user_id"):
        return api_response(success=False, message="不能删除自己", code=400)

    db = next(get_db())
    try:
        user = db.query(User).filter(User.id == user_id).first()
        if not user:
            return api_response(success=False, message="用户不存在", code=404)
        db.delete(user)
        db.commit()
        return api_response(message="用户已删除")
    except Exception as e:
        db.rollback()
        return api_response(success=False, message=str(e), code=500)
    finally:
        db.close()


@app.route("/api/users/<int:user_id>/password", methods=["PUT"])
@admin_required
def reset_password(user_id):
    data = request.get_json(force=True) or {}
    new_password = data.get("password", "").strip()
    if not new_password:
        return api_response(success=False, message="新密码不能为空", code=400)

    db = next(get_db())
    try:
        user = db.query(User).filter(User.id == user_id).first()
        if not user:
            return api_response(success=False, message="用户不存在", code=404)
        user.password_hash = generate_password_hash(new_password, method='pbkdf2:sha256')
        db.commit()
        return api_response(message="密码重置成功")
    except Exception as e:
        db.rollback()
        return api_response(success=False, message=str(e), code=500)
    finally:
        db.close()


# ===========================================================================
# 批次管理 API
# ===========================================================================

@app.route("/api/batches/generate-no", methods=["GET"])
@login_required
def generate_batch_no():
    """生成批次号，格式 {前缀}{YYYYMMDD}{序号}，前缀从系统配置 batch_code_prefix 读取"""
    db = next(get_db())
    try:
        prefix_config = db.query(SystemConfig).filter(
            SystemConfig.config_key == "batch_code_prefix"
        ).first()
        prefix = prefix_config.config_value.strip() if prefix_config and prefix_config.config_value.strip() else "BT"
        today_str = _now_cn().strftime("%Y%m%d")
        date_prefix = f"{prefix}{today_str}"
        last = db.query(Batch).filter(Batch.batch_no.like(f"{date_prefix}%")).order_by(desc(Batch.batch_no)).first()
        if last and last.batch_no.startswith(date_prefix):
            try:
                last_seq = int(last.batch_no[-3:])
                seq = last_seq + 1
            except ValueError:
                seq = 1
        else:
            seq = 1
        return api_response(data={"batch_no": f"{date_prefix}{seq:03d}"})
    finally:
        db.close()


@app.route("/api/batches", methods=["POST"])
@login_required
def create_batch():
    data = request.get_json(force=True) or {}

    # Map frontend field names → backend model field names
    if "name" in data and "description" not in data:
        data["description"] = data["name"]
    if "deadline" in data and "estimated_completion" not in data:
        data["estimated_completion"] = data["deadline"]
    if "start_date" in data and "started_at" not in data:
        data["started_at"] = data["start_date"]
    if "remark" in data and "completion_note" not in data:
        data["completion_note"] = data["remark"]

    batch_no = (data.get("batch_no") or "").strip()
    description = (data.get("description") or "").strip()
    operator = (data.get("operator") or session.get("display_name") or "").strip()

    db = next(get_db())
    try:
        # 若未手动指定批次号，自动生成
        if not batch_no:
            prefix_config = db.query(SystemConfig).filter(
                SystemConfig.config_key == "batch_code_prefix"
            ).first()
            prefix = prefix_config.config_value.strip() if prefix_config and prefix_config.config_value.strip() else "BT"
            today_str = _now_cn().strftime("%Y%m%d")
            date_prefix = f"{prefix}{today_str}"
            last = db.query(Batch).filter(Batch.batch_no.like(f"{date_prefix}%")).order_by(desc(Batch.batch_no)).first()
            if last and last.batch_no.startswith(date_prefix):
                try:
                    last_seq = int(last.batch_no[-3:])
                    seq = last_seq + 1
                except ValueError:
                    seq = 1
            else:
                seq = 1
            batch_no = f"{date_prefix}{seq:03d}"
    finally:
        db.close()

    if not operator:
        return api_response(success=False, message="操作员不能为空", code=400)

    # 处理新增字段
    total_count = data.get("total_count")
    try:
        total_count = int(total_count) if total_count is not None else 0
    except (ValueError, TypeError):
        total_count = 0

    arrival_date = None
    if data.get("arrival_date"):
        try:
            arrival_date = datetime.fromisoformat(data["arrival_date"])
        except (ValueError, TypeError):
            pass

    courier = (data.get("courier") or "").strip()
    tracking_no = (data.get("tracking_no") or "").strip()
    test_date = None
    if data.get("test_date"):
        try:
            test_date = datetime.fromisoformat(data["test_date"])
        except (ValueError, TypeError):
            pass
    test_standard = (data.get("test_standard") or "").strip()

    estimated_completion = None
    if data.get("estimated_completion"):
        try:
            estimated_completion = datetime.strptime(data["estimated_completion"], "%Y-%m-%d").date()
        except (ValueError, TypeError):
            pass

    actual_completion = None
    if data.get("actual_completion"):
        try:
            actual_completion = datetime.strptime(data["actual_completion"], "%Y-%m-%d").date()
        except (ValueError, TypeError):
            pass

    completion_note = (data.get("completion_note") or "").strip()

    db = next(get_db())
    try:
        # Resolve standard_id → standard name
        if not test_standard and "standard_id" in data:
            try:
                std_id = int(data["standard_id"])
                std = db.query(TestStandard).filter(TestStandard.id == std_id).first()
                if std:
                    test_standard = std.name
            except (ValueError, TypeError):
                pass

        if db.query(Batch).filter(Batch.batch_no == batch_no).first():
            return api_response(success=False, message="批次号已存在", code=409)

        batch = Batch(
            batch_no=batch_no,
            description=description,
            operator=operator,
            status="planned",
            device_count=0,
            total_count=total_count,
            arrival_date=arrival_date,
            courier=courier,
            tracking_no=tracking_no,
            test_date=test_date,
            test_standard=test_standard,
            estimated_completion=estimated_completion,
            actual_completion=actual_completion,
            completion_note=completion_note,
        )
        db.add(batch)
        db.commit()
        db.refresh(batch)
        return api_response(data=batch.to_dict(), message="批次创建成功", code=201)
    except Exception as e:
        db.rollback()
        return api_response(success=False, message=str(e), code=500)
    finally:
        db.close()


@app.route("/api/batches", methods=["GET"])
@login_required
def list_batches():
    db = next(get_db())
    try:
        page = request.args.get("page", 1, type=int)
        per_page = request.args.get("per_page", 20, type=int)
        search = request.args.get("search", "").strip()
        status_filter = request.args.get("status", "").strip()

        query = db.query(Batch)

        if search:
            like = f"%{search}%"
            query = query.filter(
                or_(Batch.batch_no.like(like), Batch.description.like(like))
            )
        if status_filter:
            query = query.filter(Batch.status == status_filter)

        total = query.count()
        batches = query.order_by(desc(Batch.created_at)).offset((page - 1) * per_page).limit(per_page).all()

        # Sync device_count & compute fault_count / yield_rate
        batch_dicts = []
        for b in batches:
            actual_count = db.query(func.count(DeviceTest.id)).filter(DeviceTest.batch_id == b.id).scalar()
            if actual_count != b.device_count:
                b.device_count = actual_count or 0

            normal_count = db.query(func.count(DeviceTest.id)).filter(DeviceTest.batch_id == b.id, DeviceTest.status == 'normal').scalar() or 0
            fault_count = db.query(func.count(DeviceTest.id)).filter(DeviceTest.batch_id == b.id, DeviceTest.status == 'fault').scalar() or 0
            pending_count = db.query(func.count(DeviceTest.id)).filter(DeviceTest.batch_id == b.id, DeviceTest.status == 'pending').scalar() or 0
            yield_rate = round(normal_count / b.total_count * 100, 1) if b.total_count > 0 else 0

            d = b.to_dict()
            d["normal_count"] = normal_count
            d["fault_count"] = fault_count
            d["pending_count"] = pending_count
            d["yield_rate"] = yield_rate
            batch_dicts.append(d)

        db.commit()

        return api_response(
            data={
                "items": batch_dicts,
                "total": total,
                "page": page,
                "per_page": per_page,
                "total_pages": max(1, (total + per_page - 1) // per_page),
            }
        )
    except Exception as e:
        return api_response(success=False, message=str(e), code=500)
    finally:
        db.close()


@app.route("/api/batches/<int:batch_id>", methods=["GET"])
@login_required
def get_batch(batch_id):
    db = next(get_db())
    try:
        batch = db.query(Batch).filter(Batch.id == batch_id).first()
        if not batch:
            return api_response(success=False, message="批次不存在", code=404)

        actual_count = db.query(func.count(DeviceTest.id)).filter(DeviceTest.batch_id == batch_id).scalar()
        normal_count = db.query(func.count(DeviceTest.id)).filter(DeviceTest.batch_id == batch_id, DeviceTest.status == 'normal').scalar()
        fault_count = db.query(func.count(DeviceTest.id)).filter(DeviceTest.batch_id == batch_id, DeviceTest.status == 'fault').scalar()
        pending_count = db.query(func.count(DeviceTest.id)).filter(DeviceTest.batch_id == batch_id, DeviceTest.status == 'pending').scalar()
        result = batch.to_dict()
        result["device_count"] = actual_count or 0
        result["normal_count"] = normal_count or 0
        result["fault_count"] = fault_count or 0
        result["pending_count"] = pending_count or 0

        # Collect unique operators from device tests in this batch
        operators_query = db.query(DeviceTest.operator).filter(
            DeviceTest.batch_id == batch_id,
            DeviceTest.operator.isnot(None),
            DeviceTest.operator != ''
        ).distinct().all()
        result["operators"] = [op[0] for op in operators_query if op[0]]

        # Attach metrics from TestStandard if test_standard name matches
        std_name = (batch.test_standard or "").strip()
        if std_name:
            standard = db.query(TestStandard).filter(TestStandard.name == std_name).first()
            if standard:
                result["metrics"] = json.loads(standard.metrics) if standard.metrics else []
                result["standard_name"] = standard.name
            else:
                result["metrics"] = []
                result["standard_name"] = std_name
        else:
            result["metrics"] = []
            result["standard_name"] = ""

        return api_response(data=result)
    finally:
        db.close()


@app.route("/api/batches/<int:batch_id>", methods=["PUT"])
@login_required
def update_batch(batch_id):
    data = request.get_json(force=True) or {}

    # Map frontend field names → backend model field names
    if "name" in data and "description" not in data:
        data["description"] = data["name"]
    if "deadline" in data and "estimated_completion" not in data:
        data["estimated_completion"] = data["deadline"]
    if "start_date" in data and "started_at" not in data:
        data["started_at"] = data["start_date"]
    if "remark" in data and "completion_note" not in data:
        data["completion_note"] = data["remark"]

    db = next(get_db())
    try:
        batch = db.query(Batch).filter(Batch.id == batch_id).first()
        if not batch:
            return api_response(success=False, message="批次不存在", code=404)

        if "description" in data:
            batch.description = (data["description"] or "").strip()
        if "status" in data and data["status"] in ("planned", "in_progress", "paused", "completed", "cancelled"):
            batch.status = data["status"]
            if data["status"] == "in_progress":
                batch.actual_completion = None
        if "total_count" in data:
            try:
                batch.total_count = int(data["total_count"])
            except (ValueError, TypeError):
                pass
        if "arrival_date" in data:
            try:
                batch.arrival_date = datetime.fromisoformat(data["arrival_date"]) if data["arrival_date"] else None
            except (ValueError, TypeError):
                pass
        if "courier" in data:
            batch.courier = (data["courier"] or "").strip()
        if "tracking_no" in data:
            batch.tracking_no = (data["tracking_no"] or "").strip()
        if "test_date" in data:
            try:
                batch.test_date = datetime.fromisoformat(data["test_date"]) if data["test_date"] else None
            except (ValueError, TypeError):
                pass
        if "test_standard" in data:
            batch.test_standard = (data["test_standard"] or "").strip()
        if "estimated_completion" in data:
            val = data.get("estimated_completion")
            if val:
                try:
                    batch.estimated_completion = datetime.strptime(val, "%Y-%m-%d").date()
                except (ValueError, TypeError):
                    pass
            else:
                batch.estimated_completion = None
        if "actual_completion" in data:
            val = data.get("actual_completion")
            if val:
                try:
                    batch.actual_completion = datetime.strptime(val, "%Y-%m-%d").date()
                except (ValueError, TypeError):
                    pass
            else:
                batch.actual_completion = None
        if "completion_note" in data:
            batch.completion_note = (data["completion_note"] or "").strip()
        if "metrics" in data:
            batch.metrics = json.dumps(data["metrics"], ensure_ascii=False)

        batch.updated_at = _now_cn()
        db.commit()
        db.refresh(batch)
        return api_response(data=batch.to_dict(), message="批次更新成功")
    except Exception as e:
        db.rollback()
        return api_response(success=False, message=str(e), code=500)
    finally:
        db.close()


@app.route("/api/batches/<int:batch_id>", methods=["DELETE"])
@login_required
def delete_batch(batch_id):
    db = next(get_db())
    try:
        batch = db.query(Batch).filter(Batch.id == batch_id).first()
        if not batch:
            return api_response(success=False, message="批次不存在", code=404)

        # 解绑设备（置 NULL 而非删除）
        db.query(DeviceTest).filter(DeviceTest.batch_id == batch_id).update({"batch_id": None})
        db.delete(batch)
        db.commit()
        return api_response(message="批次已删除，关联设备已解绑")
    except Exception as e:
        db.rollback()
        return api_response(success=False, message=str(e), code=500)
    finally:
        db.close()


@app.route("/api/batches/<int:batch_id>/status", methods=["PUT"])
@login_required
def update_batch_status(batch_id):
    """更新批次状态：planned / in_progress / paused / completed / cancelled"""
    data = request.get_json(force=True) or {}
    new_status = data.get("status", "").strip()
    valid_statuses = ("planned", "in_progress", "paused", "completed", "cancelled")
    if new_status not in valid_statuses:
        return api_response(success=False, message=f"无效状态，仅允许: {', '.join(valid_statuses)}", code=400)

    db = next(get_db())
    try:
        batch = db.query(Batch).filter(Batch.id == batch_id).first()
        if not batch:
            return api_response(success=False, message="批次不存在", code=404)

        # 暂停/取消：原因必填
        if new_status in ("paused", "cancelled"):
            reason = data.get("status_reason", "").strip()
            if not reason:
                return api_response(success=False, message=f"状态为「{new_status}」时必须填写原因", code=400)
            batch.status_reason = reason

        # 进行中：自动填入 started_at
        if new_status == "in_progress" and not batch.started_at:
            started_at = data.get("started_at")
            if started_at:
                try:
                    batch.started_at = datetime.fromisoformat(started_at)
                except (ValueError, TypeError):
                    batch.started_at = _now_cn()
            else:
                batch.started_at = _now_cn()

        # 完成：自动填入 actual_completion
        if new_status == "completed":
            now = _now_cn()
            actual_completion = data.get("actual_completion")
            if actual_completion:
                try:
                    batch.actual_completion = datetime.strptime(actual_completion, "%Y-%m-%d").date()
                except (ValueError, TypeError):
                    batch.actual_completion = now.date()
            else:
                batch.actual_completion = now.date()
            # 也更新 started_at（如果还没设置）
            if not batch.started_at:
                batch.started_at = now

        batch.status = new_status
        batch.updated_at = _now_cn()
        db.commit()
        db.refresh(batch)
        return api_response(data=batch.to_dict(), message=f"批次状态已更新为「{new_status}」")
    except Exception as e:
        db.rollback()
        return api_response(success=False, message=str(e), code=500)
    finally:
        db.close()


# --- 批次内设备 ---

@app.route("/api/batches/<int:batch_id>/devices", methods=["GET"])
@login_required
def list_batch_devices(batch_id):
    db = next(get_db())
    try:
        batch = db.query(Batch).filter(Batch.id == batch_id).first()
        if not batch:
            return api_response(success=False, message="批次不存在", code=404)

        page = request.args.get("page", 1, type=int)
        per_page = request.args.get("per_page", 20, type=int)
        search = request.args.get("search", "").strip()
        status_filter = request.args.get("status", "").strip()
        sort_by = request.args.get("sort_by", "").strip()
        sort_order = request.args.get("sort_order", "desc").strip()
        # 新增：不限分页，获取全量数据（用于统计/完成弹窗）
        no_page = request.args.get("no_page", "").strip().lower() in ("1", "true", "yes")

        query = db.query(DeviceTest).options(joinedload(DeviceTest.batch)).filter(DeviceTest.batch_id == batch_id)

        if search:
            like = f"%{search}%"
            query = query.filter(
                or_(
                    DeviceTest.board_mac.like(like),
                    DeviceTest.wireless_mac.like(like),
                    DeviceTest.ip_address.like(like),
                    DeviceTest.operator.like(like),
                    DeviceTest.device_code.like(like),
                )
            )
        if status_filter:
            query = query.filter(DeviceTest.status == status_filter)

        total = query.count()

        # 排序：有 sort_by 时按指定字段排序，否则按 sort_order 升序，null 排最后
        if sort_by and hasattr(DeviceTest, sort_by):
            sort_col = getattr(DeviceTest, sort_by)
            if sort_order == "asc":
                query = query.order_by(sort_col.asc())
            else:
                query = query.order_by(sort_col.desc())
        else:
            query = query.order_by(DeviceTest.sort_order.is_(None), DeviceTest.sort_order.asc(), DeviceTest.id.asc())

        total_pages = max(1, (total + per_page - 1) // per_page)

        if no_page:
            # 返回全量，不分页
            all_devices = query.all()
            return api_response(
                data={
                    "items": [d.to_dict() for d in all_devices],
                    "total": total,
                    "page": 1,
                    "per_page": total,
                    "total_pages": 1,
                }
            )
        else:
            # 用 SQL LIMIT/OFFSET 分页，避免全部加载到内存
            devices = query.offset((page - 1) * per_page).limit(per_page).all()
            return api_response(
                data={
                    "items": [d.to_dict() for d in devices],
                    "total": total,
                    "page": page,
                    "per_page": per_page,
                    "total_pages": total_pages,
                }
            )
    except Exception as e:
        return api_response(success=False, message=str(e), code=500)
    finally:
        db.close()


@app.route("/api/batches/<int:batch_id>/metrics", methods=["GET"])
@login_required
def get_batch_metrics(batch_id):
    """获取批次完整指标数据：关联标准、全量设备结果。
    前端一次请求替代 N+2 次 HTTP 调用，彻底消除翻页卡顿。"""
    db = next(get_db())
    try:
        batch = db.query(Batch).filter(Batch.id == batch_id).first()
        if not batch:
            return api_response(success=False, message="批次不存在", code=404)

        batch_std_name = (batch.test_standard or "").strip()

        # 1. 获取批次关联的测试标准
        all_standards = db.query(TestStandard).all()
        metric_standards = [s.to_dict() for s in all_standards] if not batch_std_name else [s.to_dict() for s in all_standards if s.name == batch_std_name]

        # 2. 获取批次所有设备 ID
        device_ids = [d[0] for d in db.query(DeviceTest.id).filter(DeviceTest.batch_id == batch_id).all()]

        # 3. 批量获取所有测试结果
        if device_ids:
            results = db.query(DeviceTestResult).filter(
                DeviceTestResult.device_id.in_(device_ids)
            ).order_by(DeviceTestResult.standard_id, DeviceTestResult.id).all()
        else:
            results = []

        # 按 device_id 分组（已为扁平 dict 列表，无需二次分组）
        grouped = {}
        for r in results:
            did = r.device_id
            if did not in grouped:
                grouped[did] = []
            grouped[did].append(r.to_dict())

        # 确保返回中每个设备 ID 都有条目
        for did in device_ids:
            if did not in grouped:
                grouped[did] = []

        return api_response(data={
            "standard_name": batch_std_name,
            "metric_standards": metric_standards,
            "device_ids": device_ids,
            "results": grouped,
        })
    except Exception as e:
        return api_response(success=False, message=str(e), code=500)
    finally:
        db.close()


@app.route("/api/batches/<int:batch_id>/init-results", methods=["POST"])
@login_required
def init_batch_results(batch_id):
    """为批次中所有设备的每个标准+指标，创建默认 pass=0 的记录（如缺失）"""
    db = next(get_db())
    try:
        batch = db.query(Batch).filter(Batch.id == batch_id).first()
        if not batch:
            return api_response(success=False, message="批次不存在", code=404)

        # 获取批次关联的标准名称
        batch_std_name = (batch.test_standard or "").strip()
        all_standards = db.query(TestStandard).all()
        if batch_std_name:
            standards = [s for s in all_standards if s.name == batch_std_name]
        else:
            standards = all_standards

        if not standards:
            return api_response(success=False, message="未找到关联的测试标准", code=400)

        # 获取批次所有设备
        devices = db.query(DeviceTest).filter(DeviceTest.batch_id == batch_id).all()
        if not devices:
            return api_response(success=True, data={"created": 0, "skipped": 0, "total": 0})

        created = 0
        skipped = 0

        for device in devices:
            for std in standards:
                metrics = json.loads(std.metrics) if std.metrics else []
                for m in metrics:
                    metric_name = m.get("name", "")
                    if not metric_name:
                        continue
                    # 检查是否已存在
                    existing = db.query(DeviceTestResult).filter(
                        DeviceTestResult.device_id == device.id,
                        DeviceTestResult.standard_id == std.id,
                        DeviceTestResult.metric_name == metric_name,
                    ).first()
                    if existing:
                        skipped += 1
                    else:
                        r = DeviceTestResult(
                            device_id=device.id,
                            standard_id=std.id,
                            metric_name=metric_name,
                            metric_value="",
                            expected_value=m.get("value", ""),
                            pass_=0,
                            notes="",
                        )
                        db.add(r)
                        created += 1

        db.commit()
        return api_response(
            data={"created": created, "skipped": skipped, "total": created + skipped},
            message=f"已创建 {created} 条缺失记录，跳过 {skipped} 条已有记录"
        )
    except Exception as e:
        db.rollback()
        return api_response(success=False, message=str(e), code=500)
    finally:
        db.close()


@app.route("/api/batches/<int:batch_id>/devices", methods=["POST"])
@login_required
def create_batch_device(batch_id):
    db = next(get_db())
    try:
        batch = db.query(Batch).filter(Batch.id == batch_id).first()
        if not batch:
            return api_response(success=False, message="批次不存在", code=404)

        data = request.get_json(force=True)
        errors = validate_device_data(data)
        if errors:
            return api_response(success=False, message="; ".join(errors), code=400)

        # 序号：前端传入则用传入值，否则取当前批次最大序号+1
        if "sort_order" in data and data["sort_order"] is not None:
            sort_order = int(data["sort_order"])
        else:
            max_seq = db.query(func.max(DeviceTest.sort_order)).filter(DeviceTest.batch_id == batch_id).scalar()
            sort_order = (max_seq + 1) if max_seq is not None else 1

        # 批次内唯一性检查：同批次中板卡MAC+无线MAC不可重复（仅当无线MAC非空时）
        board_mac = normalize_mac(data["board_mac"])
        wireless_mac_raw = data["wireless_mac"].strip()
        wireless_mac = normalize_mac(wireless_mac_raw) if wireless_mac_raw else None
        ip_addr = data["ip_address"].strip() or None
        if wireless_mac:
            existing = db.query(DeviceTest).filter(
                DeviceTest.batch_id == batch_id,
                DeviceTest.board_mac == board_mac,
                DeviceTest.wireless_mac == wireless_mac,
            ).first()
            if existing:
                return api_response(success=False, message="该板卡MAC+无线MAC组合在当前批次中已存在", code=409)

        status_val = data.get("status", "normal")
        fault_disposition_val = (data.get("fault_disposition") or 'pending') if status_val == "fault" else (data.get("fault_disposition") or '')
        user_device_code = (data.get("device_code") or "").strip()
        if user_device_code and len(user_device_code) == 8 and user_device_code.isalnum():
            device_code_val = user_device_code.upper()
        else:
            device_code_val = board_mac.replace(":", "").replace("-", "").upper()
        device = DeviceTest(
            board_mac=board_mac,
            device_code=device_code_val,
            device_type=(data.get("device_type") or "").strip(),
            wireless_mac=wireless_mac,
            ip_address=ip_addr,
            status=status_val,
            fault_reason=(data.get("fault_reason") or "").strip() or None,
            fault_disposition=fault_disposition_val,
            return_date=datetime.strptime(data["return_date"], "%Y-%m-%d").date() if data.get("return_date") else None,
            return_tracking=(data.get("return_tracking") or "").strip() or None,
            test_date=parse_iso_date(data.get("test_date")) or _now_cn(),
            operator=(data.get("operator") or "").strip() or None,
            notes=(data.get("notes") or "").strip() or None,
            sort_order=sort_order,
            batch_id=batch_id,
        )
        db.add(device)

        # 更新批次设备计数
        count = db.query(func.count(DeviceTest.id)).filter(DeviceTest.batch_id == batch_id).scalar() or 0
        batch.device_count = count + 1

        db.commit()
        db.refresh(device)
        return api_response(data=device.to_dict(), message="设备添加成功", code=201)
    except IntegrityError as e:
        db.rollback()
        app.logger.error(f"IntegrityError for device {data}: {e}")
        return api_response(success=False, message=f"数据库约束冲突: {e}", code=409)
    except Exception as e:
        db.rollback()
        return api_response(success=False, message=str(e), code=500)
    finally:
        db.close()


# ===========================================================================
# 设备测试结果 API
# ===========================================================================

@app.route("/api/devices/<int:device_id>/results", methods=["GET"])
@login_required
def get_device_results(device_id):
    """获取设备的所有测试结果（按 standard_id 分组）"""
    db = next(get_db())
    try:
        results = db.query(DeviceTestResult).filter(
            DeviceTestResult.device_id == device_id
        ).order_by(DeviceTestResult.standard_id, DeviceTestResult.id).all()

        grouped = {}
        for r in results:
            sid = r.standard_id
            if sid not in grouped:
                grouped[sid] = {"standard_id": sid, "items": []}
            grouped[sid]["items"].append(r.to_dict())

        return api_response(data=list(grouped.values()))
    finally:
        db.close()


@app.route("/api/devices/<int:device_id>/results", methods=["PUT"])
@login_required
def update_device_results(device_id):
    """批量更新设备测试结果：先删旧数据，再插入新数据"""
    data = request.get_json(force=True) or {}
    results_data = data.get("results", [])

    db = next(get_db())
    try:
        device = db.query(DeviceTest).filter(DeviceTest.id == device_id).first()
        if not device:
            return api_response(success=False, message="设备不存在", code=404)

        # 删除该设备旧结果
        db.query(DeviceTestResult).filter(DeviceTestResult.device_id == device_id).delete()

        # 批量插入新结果
        for item in results_data:
            result = DeviceTestResult(
                device_id=device_id,
                standard_id=item.get("standard_id", 0),
                metric_name=item.get("metric_name", ""),
                metric_value=item.get("metric_value", ""),
                expected_value=item.get("expected_value", ""),
                pass_=1 if item.get("pass") else 0,
                notes=item.get("notes", ""),
            )
            db.add(result)

        db.commit()

        # 返回更新后的结果
        results = db.query(DeviceTestResult).filter(
            DeviceTestResult.device_id == device_id
        ).order_by(DeviceTestResult.standard_id, DeviceTestResult.id).all()
        grouped = {}
        for r in results:
            sid = r.standard_id
            if sid not in grouped:
                grouped[sid] = {"standard_id": sid, "items": []}
            grouped[sid]["items"].append(r.to_dict())

        return api_response(data=list(grouped.values()), message="测试结果已更新")
    except Exception as e:
        db.rollback()
        return api_response(success=False, message=str(e), code=500)
    finally:
        db.close()


# ===========================================================================
# 测试标准管理 API
# ===========================================================================

@app.route("/api/standards", methods=["GET"])
@login_required
def list_standards():
    db = next(get_db())
    try:
        standards = db.query(TestStandard).order_by(TestStandard.id).all()
        return api_response(data=[s.to_dict() for s in standards])
    finally:
        db.close()


@app.route("/api/standards", methods=["POST"])
@login_required
def create_standard():
    data = request.get_json(force=True) or {}
    name = data.get("name", "").strip()
    description = data.get("description", "").strip()
    metrics = data.get("metrics", [])

    if not name:
        return api_response(success=False, message="标准名称不能为空", code=400)

    db = next(get_db())
    try:
        metrics_json = json.dumps(metrics, ensure_ascii=False)
        standard = TestStandard(name=name, description=description, metrics=metrics_json)
        db.add(standard)
        db.commit()
        db.refresh(standard)
        return api_response(data=standard.to_dict(), message="标准创建成功", code=201)
    except IntegrityError:
        db.rollback()
        return api_response(success=False, message="标准名称已存在", code=409)
    except Exception as e:
        db.rollback()
        return api_response(success=False, message=str(e), code=500)
    finally:
        db.close()


@app.route("/api/standards/<int:standard_id>", methods=["GET"])
@login_required
def get_standard(standard_id):
    db = next(get_db())
    try:
        s = db.query(TestStandard).filter(TestStandard.id == standard_id).first()
        if not s:
            return api_response(success=False, message="标准不存在", code=404)
        return api_response(data=s.to_dict())
    finally:
        db.close()


@app.route("/api/standards/<int:standard_id>", methods=["PUT"])
@login_required
def update_standard(standard_id):
    data = request.get_json(force=True) or {}
    db = next(get_db())
    try:
        s = db.query(TestStandard).filter(TestStandard.id == standard_id).first()
        if not s:
            return api_response(success=False, message="标准不存在", code=404)
        if "name" in data:
            s.name = data["name"].strip()
        if "description" in data:
            s.description = data["description"].strip()
        if "metrics" in data:
            s.metrics = json.dumps(data["metrics"], ensure_ascii=False)
        s.updated_at = _now_cn()
        db.commit()
        db.refresh(s)
        return api_response(data=s.to_dict(), message="标准更新成功")
    except IntegrityError:
        db.rollback()
        return api_response(success=False, message="标准名称已存在", code=409)
    except Exception as e:
        db.rollback()
        return api_response(success=False, message=str(e), code=500)
    finally:
        db.close()


@app.route("/api/standards/<int:standard_id>", methods=["DELETE"])
@login_required
def delete_standard(standard_id):
    db = next(get_db())
    try:
        s = db.query(TestStandard).filter(TestStandard.id == standard_id).first()
        if not s:
            return api_response(success=False, message="标准不存在", code=404)
        db.delete(s)
        db.commit()
        return api_response(message="标准已删除")
    except Exception as e:
        db.rollback()
        return api_response(success=False, message=str(e), code=500)
    finally:
        db.close()


# ===========================================================================
# 静态文件服务
# ===========================================================================
FRONTEND_DIR = os.path.join(PROJECT_DIR, "frontend")


@app.route("/")
@app.route("/frontend/")
@app.route("/frontend/index.html")
def serve_index():
    resp = send_from_directory(FRONTEND_DIR, "index.html")
    resp.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    resp.headers["Pragma"] = "no-cache"
    resp.headers["Expires"] = "0"
    return resp


@app.route("/<path:filename>")
def serve_static(filename):
    if filename.startswith("frontend/"):
        filename = filename[len("frontend/"):]
    return send_from_directory(FRONTEND_DIR, filename)


# ===========================================================================
# 设备测试记录 CRUD
# ===========================================================================

@app.route("/api/health", methods=["GET"])
def health_check():
    try:
        db = next(get_db())
        db.execute(text("SELECT 1"))
        return api_response(data={"db": "connected", "status": "ok"})
    except Exception as e:
        return api_response(success=False, message=f"数据库连接失败: {str(e)}", code=500)


@app.route("/api/timezone", methods=["GET"])
def get_timezone():
    """返回当前系统时区配置（无需登录）"""
    tz = _get_system_tz()
    return api_response(data={"timezone": _CN_TZ_CACHE["name"]})


# ===== 常用命令 API =====
@app.route("/api/commands", methods=["GET"])
@login_required
def list_commands():
    db = next(get_db())
    try:
        search = request.args.get("search", "").strip()
        category = request.args.get("category", "").strip()
        q = db.query(TestCommand)
        if search:
            like = f"%{search}%"
            q = q.filter(or_(TestCommand.title.like(like), TestCommand.command.like(like), TestCommand.notes.like(like)))
        if category:
            q = q.filter(TestCommand.category == category)
        q = q.order_by(TestCommand.sort_order.asc(), TestCommand.id.asc())
        commands = q.all()

        # 查询当前用户置顶列表
        user_id = session["user_id"]
        pinned_ids = set(
            p.command_id for p in db.query(UserCommandPin.command_id)
            .filter(UserCommandPin.user_id == user_id).all()
        )
        result = []
        for c in commands:
            d = c.to_dict()
            d["is_pinned"] = c.id in pinned_ids
            result.append(d)
        return api_response(data=result)
    except Exception as e:
        return api_response(success=False, message=str(e), code=500)
    finally:
        db.close()


@app.route("/api/commands", methods=["POST"])
@login_required
def create_command():
    db = next(get_db())
    try:
        data = request.get_json(force=True)
        title = (data.get("title") or "").strip()
        command = (data.get("command") or "").strip()
        if not title or not command:
            return api_response(success=False, message="标题和命令内容不能为空", code=400)
        cat = (data.get("category") or "其他").strip()[:32]
        notes = (data.get("notes") or "").strip()
        sort = data.get("sort_order", 0)
        try:
            sort = int(sort)
        except Exception:
            sort = 0
        tc = TestCommand(title=title, command=command, category=cat, notes=notes, sort_order=sort)
        db.add(tc)
        db.commit()
        return api_response(data=tc.to_dict(), message="创建成功")
    except Exception as e:
        db.rollback()
        return api_response(success=False, message=str(e), code=500)
    finally:
        db.close()


@app.route("/api/commands/<int:cmd_id>", methods=["PUT"])
@login_required
def update_command(cmd_id):
    db = next(get_db())
    try:
        tc = db.query(TestCommand).filter(TestCommand.id == cmd_id).first()
        if not tc:
            return api_response(success=False, message="命令不存在", code=404)
        data = request.get_json(force=True)
        if "title" in data:
            v = (data["title"] or "").strip()
            if not v:
                return api_response(success=False, message="标题不能为空", code=400)
            tc.title = v
        if "command" in data:
            v = (data["command"] or "").strip()
            if not v:
                return api_response(success=False, message="命令内容不能为空", code=400)
            tc.command = v
        if "category" in data:
            tc.category = (data["category"] or "其他").strip()[:32]
        if "notes" in data:
            tc.notes = (data["notes"] or "").strip()
        if "sort_order" in data:
            try:
                tc.sort_order = int(data["sort_order"])
            except Exception:
                pass
        tc.updated_at = _now_cn()
        db.commit()
        return api_response(data=tc.to_dict(), message="更新成功")
    except Exception as e:
        db.rollback()
        return api_response(success=False, message=str(e), code=500)
    finally:
        db.close()


@app.route("/api/commands/<int:cmd_id>", methods=["DELETE"])
@login_required
def delete_command(cmd_id):
    db = next(get_db())
    try:
        tc = db.query(TestCommand).filter(TestCommand.id == cmd_id).first()
        if not tc:
            return api_response(success=False, message="命令不存在", code=404)
        db.delete(tc)
        db.commit()
        return api_response(message="删除成功")
    except Exception as e:
        db.rollback()
        return api_response(success=False, message=str(e), code=500)
    finally:
        db.close()


# ===== 命令置顶 API =====

@app.route("/api/commands/pin", methods=["POST"])
@login_required
def pin_command():
    """置顶一条命令（当前用户专属）"""
    data = request.get_json(force=True)
    command_id = data.get("command_id")
    if not command_id:
        return api_response(success=False, message="缺少 command_id", code=400)
    db = next(get_db())
    try:
        user_id = session["user_id"]
        existing = db.query(UserCommandPin).filter_by(user_id=user_id, command_id=command_id).first()
        if existing:
            return api_response(data={"pinned": True}, message="已置顶")
        max_order = db.query(func.coalesce(func.max(UserCommandPin.sort_order), 0)).filter_by(user_id=user_id).scalar() or 0
        pin = UserCommandPin(user_id=user_id, command_id=command_id, sort_order=max_order + 1)
        db.add(pin)
        db.commit()
        return api_response(data={"pinned": True}, message="置顶成功", code=201)
    except Exception as e:
        db.rollback()
        return api_response(success=False, message=str(e), code=500)
    finally:
        db.close()


@app.route("/api/commands/pin/<int:command_id>", methods=["DELETE"])
@login_required
def unpin_command(command_id):
    """取消置顶"""
    db = next(get_db())
    try:
        user_id = session["user_id"]
        pin = db.query(UserCommandPin).filter_by(user_id=user_id, command_id=command_id).first()
        if pin:
            db.delete(pin)
            db.commit()
        return api_response(data={"pinned": False}, message="已取消置顶")
    except Exception as e:
        db.rollback()
        return api_response(success=False, message=str(e), code=500)
    finally:
        db.close()


@app.route("/api/commands/pins", methods=["GET"])
@login_required
def get_user_pins():
    """获取当前用户的置顶命令ID列表（按 sort_order 排序）"""
    db = next(get_db())
    try:
        user_id = session["user_id"]
        pins = db.query(UserCommandPin).filter_by(user_id=user_id).order_by(UserCommandPin.sort_order).all()
        return api_response(data=[p.command_id for p in pins])
    finally:
        db.close()


# ===== 帮助文档 API =====
@app.route("/api/help", methods=["GET"])
def help_doc():
    """返回渲染后的帮助文档 HTML"""
    readme_path = os.path.join(PROJECT_DIR, "README.md")
    try:
        with open(readme_path, "r", encoding="utf-8") as f:
            md_content = f.read()
        # 移除 YAML frontmatter（开头的 --- ... --- 块），避免 AIGC 元数据渲染到页面
        md_content = re.sub(r'^---\s*\n.*?\n---\s*\n', '', md_content, flags=re.DOTALL)
        html = md_lib.markdown(md_content, extensions=["tables", "fenced_code", "codehilite", "toc"])
        return api_response(data={"html": html})
    except FileNotFoundError:
        return api_response(success=False, message="帮助文档不存在", code=404)


@app.route("/api/devices", methods=["GET"])
@login_required
def list_devices():
    db = next(get_db())
    try:
        page = request.args.get("page", 1, type=int)
        per_page = request.args.get("per_page", 20, type=int)
        search = request.args.get("search", "").strip()
        status_filter = request.args.get("status", "").strip()
        disposition_filter = request.args.get("disposition", "").strip()
        date_from = request.args.get("date_from", "").strip()
        date_to = request.args.get("date_to", "").strip()
        batch_id = request.args.get("batch_id", "").strip()
        flow_status = request.args.get("flow_status", "").strip()
        sort_by = request.args.get("sort_by", "test_date")
        sort_order = request.args.get("sort_order", "desc")

        query = db.query(DeviceTest).options(joinedload(DeviceTest.batch))

        if search:
            like = f"%{search}%"
            query = query.filter(
                or_(
                    DeviceTest.board_mac.like(like),
                    DeviceTest.wireless_mac.like(like),
                    DeviceTest.ip_address.like(like),
                    DeviceTest.return_tracking.like(like),
                    DeviceTest.operator.like(like),
                    DeviceTest.notes.like(like),
                    DeviceTest.device_code.like(like),
                )
            )
        if status_filter:
            query = query.filter(DeviceTest.status == status_filter)
        if disposition_filter:
            if disposition_filter == "returned":
                disposition_filter = "待返厂"
            query = query.filter(DeviceTest.fault_disposition == disposition_filter)
        if date_from:
            query = query.filter(DeviceTest.test_date >= datetime.fromisoformat(date_from))
        if date_to:
            query = query.filter(DeviceTest.test_date <= datetime.fromisoformat(date_to + "T23:59:59"))
        if batch_id:
            try:
                query = query.filter(DeviceTest.batch_id == int(batch_id))
            except ValueError:
                pass
        if flow_status:
            if flow_status == "outgoing":
                query = query.filter(DeviceTest.flow_status != "在库", DeviceTest.flow_status.isnot(None), DeviceTest.flow_status != "")
            else:
                query = query.filter(DeviceTest.flow_status == flow_status)

        sort_col = getattr(DeviceTest, sort_by, DeviceTest.test_date)
        query = query.order_by(desc(sort_col) if sort_order == "desc" else asc(sort_col))

        total = query.count()
        devices = query.offset((page - 1) * per_page).limit(per_page).all()

        return api_response(
            data={
                "items": [d.to_dict() for d in devices],
                "total": total,
                "page": page,
                "per_page": per_page,
                "total_pages": max(1, (total + per_page - 1) // per_page),
            }
        )
    except Exception as e:
        return api_response(success=False, message=str(e), code=500)
    finally:
        db.close()


@app.route("/api/devices/<int:device_id>", methods=["GET"])
@login_required
def get_device(device_id):
    db = next(get_db())
    try:
        device = db.query(DeviceTest).filter(DeviceTest.id == device_id).first()
        if not device:
            return api_response(success=False, message="设备记录不存在", code=404)
        return api_response(data=device.to_dict())
    finally:
        db.close()


@app.route("/api/devices", methods=["POST"])
@login_required
def create_device():
    data = request.get_json(force=True)

    errors = validate_device_data(data)
    if errors:
        return api_response(success=False, message="; ".join(errors), code=400)

    db = next(get_db())
    try:
        batch_id = data.get("batch_id")
        if batch_id is not None:
            try:
                batch_id = int(batch_id)
                batch = db.query(Batch).filter(Batch.id == batch_id).first()
                if not batch:
                    return api_response(success=False, message="指定批次不存在", code=400)
            except (ValueError, TypeError):
                batch_id = None

        board_mac = normalize_mac(data["board_mac"])
        wireless_mac_raw = data["wireless_mac"].strip()
        wireless_mac = normalize_mac(wireless_mac_raw) if wireless_mac_raw else None
        ip_addr = data["ip_address"].strip() or None

        # 批次内唯一性检查（仅当无线MAC非空时）
        if batch_id is not None and wireless_mac:
            existing = db.query(DeviceTest).filter(
                DeviceTest.batch_id == batch_id,
                DeviceTest.board_mac == board_mac,
                DeviceTest.wireless_mac == wireless_mac,
            ).first()
            if existing:
                return api_response(success=False, message="该板卡MAC+无线MAC组合在当前批次中已存在", code=409)

        status_val = data.get("status", "normal")
        fault_disposition_val = (data.get("fault_disposition") or 'pending') if status_val == "fault" else (data.get("fault_disposition") or '')
        user_device_code = (data.get("device_code") or "").strip()
        if user_device_code and len(user_device_code) == 8 and user_device_code.isalnum():
            device_code_val = user_device_code.upper()
        else:
            device_code_val = board_mac.replace(":", "").replace("-", "").upper()
        device = DeviceTest(
            board_mac=board_mac,
            device_code=device_code_val,
            device_type=(data.get("device_type") or "").strip(),
            wireless_mac=wireless_mac,
            ip_address=ip_addr,
            status=status_val,
            fault_reason=(data.get("fault_reason") or "").strip() or None,
            fault_disposition=fault_disposition_val,
            return_date=datetime.strptime(data["return_date"], "%Y-%m-%d").date() if data.get("return_date") else None,
            return_tracking=(data.get("return_tracking") or "").strip() or None,
            test_date=parse_iso_date(data.get("test_date")) or _now_cn(),
            operator=(data.get("operator") or "").strip() or None,
            notes=(data.get("notes") or "").strip() or None,
            batch_id=batch_id,
        )
        db.add(device)

        if batch_id:
            count = db.query(func.count(DeviceTest.id)).filter(DeviceTest.batch_id == batch_id).scalar() or 0
            db.query(Batch).filter(Batch.id == batch_id).update({"device_count": count + 1})

        # 待返厂设备无需自动创建返厂跟踪记录，由返厂跟踪页面统一创建

        db.commit()
        db.refresh(device)
        return api_response(data=device.to_dict(), message="设备记录创建成功", code=201)
    except IntegrityError as e:
        db.rollback()
        app.logger.error(f"IntegrityError for device board_mac={data.get('board_mac', '?')}: {e}")
        return api_response(success=False, message=f"数据库约束冲突: {e}", code=409)
    except Exception as e:
        db.rollback()
        return api_response(success=False, message=str(e), code=500)
    finally:
        db.close()


@app.route("/api/devices/<int:device_id>", methods=["PUT", "PATCH"])
@login_required
def update_device(device_id):
    db = next(get_db())
    try:
        device = db.query(DeviceTest).filter(DeviceTest.id == device_id).first()
        if not device:
            return api_response(success=False, message="设备记录不存在", code=404)

        data = request.get_json(force=True)

        if "status" in data:
            if data["status"] == "normal":
                device.fault_reason = None
                device.fault_disposition = ''
                device.return_date = None
                device.return_tracking = None
                device.fault_return_courier = None
                device.fault_return_tracking = None

        updatable = [
            "board_mac", "wireless_mac", "ip_address", "status",
            "fault_reason", "fault_disposition", "return_date",
            "return_tracking", "fault_return_courier", "fault_return_tracking",
            "operator", "notes", "test_date", "sort_order", "device_type",
            "flow_status", "flow_purpose", "flow_destination", "flow_contact",
            "flow_phone", "flow_notes",
        ]

        old_batch_id = device.batch_id

        for key in updatable:
            if key in data:
                val = data[key]
                if key == "return_date":
                    val = datetime.strptime(val, "%Y-%m-%d").date() if val else None
                elif key == "test_date":
                    val = datetime.fromisoformat(val) if val else device.test_date
                elif key in ("board_mac", "wireless_mac"):
                    val = normalize_mac(val) if val else None
                elif isinstance(val, str):
                    val = val.strip() or None
                setattr(device, key, val)

        # 处理流转日期字段
        for date_field in ["flow_out_date", "flow_expected_return", "flow_actual_return"]:
            if date_field in data:
                val = data[date_field]
                setattr(device, date_field, datetime.strptime(val, "%Y-%m-%d").date() if val else None)

        # 标记返还时自动填写实际返还日期
        if "flow_status" in data and data["flow_status"] == "在库" and not device.flow_actual_return:
            device.flow_actual_return = _today_cn()

        # 设备编号：用户填写8位则使用，否则从板卡MAC生成
        if "device_code" in data:
            custom_code = (data["device_code"] or "").strip()
            if custom_code and len(custom_code) == 8 and custom_code.isalnum():
                device.device_code = custom_code.upper()
            elif "board_mac" in data:
                device.device_code = device.board_mac.replace(":", "").replace("-", "").upper() if device.board_mac else ""
        elif "board_mac" in data:
            device.device_code = device.board_mac.replace(":", "").replace("-", "").upper() if device.board_mac else ""

        # 支持更新 batch_id
        if "batch_id" in data:
            new_bid = data["batch_id"]
            if new_bid is not None:
                try:
                    new_bid = int(new_bid)
                except (ValueError, TypeError):
                    new_bid = None
            device.batch_id = new_bid

            # 更新批次设备计数
            if old_batch_id:
                cnt = db.query(func.count(DeviceTest.id)).filter(DeviceTest.batch_id == old_batch_id).scalar() or 0
                db.query(Batch).filter(Batch.id == old_batch_id).update({"device_count": cnt})
            if new_bid:
                cnt = db.query(func.count(DeviceTest.id)).filter(DeviceTest.batch_id == new_bid).scalar() or 0
                db.query(Batch).filter(Batch.id == new_bid).update({"device_count": cnt})

        # 待返厂设备无需自动创建返厂跟踪记录，由返厂跟踪页面统一创建

        # 批次内唯一性检查（仅当 board_mac/wireless_mac/batch_id 任一变更时）
        mac_changed = "board_mac" in data or "wireless_mac" in data
        batch_changed = "batch_id" in data
        if mac_changed or batch_changed:
            bid = device.batch_id if device.batch_id is not None else old_batch_id
            if bid is not None:
                dup = db.query(DeviceTest).filter(
                    DeviceTest.batch_id == bid,
                    DeviceTest.board_mac == device.board_mac,
                    DeviceTest.wireless_mac == device.wireless_mac,
                    DeviceTest.id != device.id,
                ).first()
                if dup:
                    db.rollback()
                    return api_response(success=False, message="该板卡MAC+无线MAC组合在当前批次中已存在", code=409)

        db.commit()
        db.refresh(device)
        return api_response(data=device.to_dict(), message="设备记录更新成功")
    except IntegrityError as e:
        db.rollback()
        app.logger.error(f"IntegrityError in update_device device_id={device_id}: {e}")
        return api_response(success=False, message=f"数据库约束冲突: {e}", code=409)
    except Exception as e:
        db.rollback()
        return api_response(success=False, message=str(e), code=500)
    finally:
        db.close()


@app.route("/api/devices/<int:device_id>", methods=["DELETE"])
@login_required
def delete_device(device_id):
    db = next(get_db())
    try:
        device = db.query(DeviceTest).filter(DeviceTest.id == device_id).first()
        if not device:
            return api_response(success=False, message="设备记录不存在", code=404)

        batch_id = device.batch_id

        # 先删除关联的测试结果和返厂记录
        db.query(DeviceTestResult).filter(DeviceTestResult.device_id == device_id).delete()
        db.query(DeviceReturnRecord).filter(DeviceReturnRecord.device_id == device_id).delete()

        db.delete(device)

        # 更新批次计数
        if batch_id:
            cnt = db.query(func.count(DeviceTest.id)).filter(DeviceTest.batch_id == batch_id).scalar() or 0
            db.query(Batch).filter(Batch.id == batch_id).update({"device_count": cnt})

        db.commit()
        return api_response(message="设备记录已删除")
    except Exception as e:
        db.rollback()
        return api_response(success=False, message=str(e), code=500)
    finally:
        db.close()


# --- 设备流转 ---

@app.route("/api/devices/<int:device_id>/flow", methods=["PUT"])
@login_required
def update_device_flow(device_id):
    """更新设备流转信息"""
    data = request.get_json(force=True)
    db = next(get_db())
    try:
        device = db.query(DeviceTest).filter(DeviceTest.id == device_id).first()
        if not device:
            return api_response(success=False, message="设备不存在", code=404)

        flow_fields = ["flow_status", "flow_purpose", "flow_destination", "flow_contact", "flow_phone", "flow_notes"]
        for f in flow_fields:
            if f in data:
                setattr(device, f, (data[f] or "").strip())

        if "flow_out_date" in data:
            device.flow_out_date = datetime.strptime(data["flow_out_date"], "%Y-%m-%d").date() if data["flow_out_date"] else None
        if "flow_expected_return" in data:
            device.flow_expected_return = datetime.strptime(data["flow_expected_return"], "%Y-%m-%d").date() if data["flow_expected_return"] else None
        if "flow_actual_return" in data:
            device.flow_actual_return = datetime.strptime(data["flow_actual_return"], "%Y-%m-%d").date() if data["flow_actual_return"] else None

        # 标记返还时自动填写实际返还日期
        if data.get("flow_status") == "在库" and not device.flow_actual_return:
            device.flow_actual_return = _today_cn()

        db.commit()
        return api_response(data=device.to_dict(), message="流转信息已更新")
    except Exception as e:
        db.rollback()
        return api_response(success=False, message=str(e), code=500)
    finally:
        db.close()


@app.route("/api/devices/batch/flow", methods=["PUT"])
@login_required
def batch_update_flow():
    """批量更新设备流转状态"""
    data = request.get_json(force=True)
    ids = data.get("ids", [])
    flow_status = data.get("flow_status", "")
    if not ids or not flow_status:
        return api_response(success=False, message="参数无效", code=400)

    db = next(get_db())
    try:
        update_data = {"flow_status": flow_status}
        if flow_status == "在库":
            update_data["flow_actual_return"] = _today_cn()
        updated = db.query(DeviceTest).filter(DeviceTest.id.in_(ids)).update(update_data, synchronize_session=False)
        db.commit()
        return api_response(data={"updated": updated}, message=f"已更新 {updated} 台设备流转状态")
    except Exception as e:
        db.rollback()
        return api_response(success=False, message=str(e), code=500)
    finally:
        db.close()


@app.route("/api/statistics/flow", methods=["GET"])
@login_required
def flow_statistics():
    """获取流转统计"""
    db = next(get_db())
    try:
        today = _today_cn()
        in_stock = db.query(func.count(DeviceTest.id)).filter(DeviceTest.flow_status == "在库").scalar() or 0
        outgoing = db.query(func.count(DeviceTest.id)).filter(
            DeviceTest.flow_status != "在库", DeviceTest.flow_status.isnot(None), DeviceTest.flow_status != ""
        ).scalar() or 0
        overdue = db.query(func.count(DeviceTest.id)).filter(
            DeviceTest.flow_expected_return < today,
            DeviceTest.flow_status != "在库",
            DeviceTest.flow_status.isnot(None), DeviceTest.flow_status != ""
        ).scalar() or 0
        month_start = today.replace(day=1)
        returned_this_month = db.query(func.count(DeviceTest.id)).filter(
            DeviceTest.flow_actual_return >= month_start,
            DeviceTest.flow_actual_return.isnot(None)
        ).scalar() or 0
        return api_response(data={
            "in_stock": in_stock,
            "outgoing": outgoing,
            "overdue": overdue,
            "returned_this_month": returned_this_month
        })
    finally:
        db.close()


# --- 批量操作 ---

@app.route("/api/devices/batch", methods=["POST"])
@login_required
def batch_create():
    data = request.get_json(force=True)
    items = data.get("items", [])
    if not items:
        return api_response(success=False, message="批量数据为空", code=400)

    batch_id = data.get("batch_id")
    if batch_id is not None:
        try:
            batch_id = int(batch_id)
        except (ValueError, TypeError):
            batch_id = None

    results = {"created": 0, "failed": [], "total": len(items)}
    db = next(get_db())
    # 批次内已存在的 MAC 集合（一次查询加载，避免循环内逐条 DB 查询）
    existing_macs = set()
    if batch_id is not None:
        rows = db.query(DeviceTest.board_mac, DeviceTest.wireless_mac).filter(
            DeviceTest.batch_id == batch_id
        ).all()
        existing_macs = {(r[0].upper(), (r[1] or "").upper()) for r in rows if r[1]}
    # 本次导入中已见过的 MAC 集合（防重复提交）
    seen_this_batch = set()
    try:
        for i, item in enumerate(items):
            errors = validate_device_data(item)
            if errors:
                results["failed"].append({"index": i, "mac": item.get("board_mac", "?"), "errors": errors})
                continue
            bm = normalize_mac(item["board_mac"])
            wm_raw = item["wireless_mac"].strip()
            wm = normalize_mac(wm_raw) if wm_raw else None
            ip_raw = item["ip_address"].strip()
            ip_addr = ip_raw if ip_raw else None
            # 批次内唯一性检查（仅当 wireless_mac 非空时）
            if batch_id is not None and wm:
                key = (bm, wm)
                if key in existing_macs or key in seen_this_batch:
                    results["failed"].append({"index": i, "mac": bm, "errors": ["该板卡MAC+无线MAC组合在当前批次中已存在"]})
                    continue
                seen_this_batch.add(key)
            try:
                item_status = item.get("status", "pending")
                fault_disposition_val = (item.get("fault_disposition") or 'pending') if item_status == "fault" else (item.get("fault_disposition") or '')
                device = DeviceTest(
                    board_mac=bm,
                    wireless_mac=wm,
                    ip_address=ip_addr,
                    status=item_status,
                    fault_reason=(item.get("fault_reason") or "").strip() or None,
                    fault_disposition=fault_disposition_val,
                    return_date=datetime.strptime(item["return_date"], "%Y-%m-%d").date() if item.get("return_date") else None,
                    return_tracking=(item.get("return_tracking") or "").strip() or None,
                    operator=(item.get("operator") or "").strip() or None,
                    notes=(item.get("notes") or "").strip() or None,
                    batch_id=batch_id,
                )
                db.add(device)
                results["created"] += 1
            except IntegrityError as e:
                db.rollback()
                app.logger.error(f"IntegrityError for device item[{i}] board_mac={item.get('board_mac', '?')}: {e}")
                results["failed"].append({"index": i, "mac": item.get("board_mac", "?"), "errors": [f"数据库约束冲突: {e}"]})
            except Exception as e:
                results["failed"].append({"index": i, "mac": item.get("board_mac", "?"), "errors": [str(e)]})

        if batch_id and results["created"] > 0:
            cnt = db.query(func.count(DeviceTest.id)).filter(DeviceTest.batch_id == batch_id).scalar() or 0
            db.query(Batch).filter(Batch.id == batch_id).update({"device_count": cnt})

        db.commit()
        return api_response(data=results, message=f"批量录入完成: {results['created']}/{results['total']}", code=201)
    except IntegrityError as e:
        db.rollback()
        app.logger.error(f"IntegrityError in batch_create commit: {e}")
        return api_response(success=False, message=f"数据库约束冲突: {e}", code=409)
    except Exception as e:
        db.rollback()
        return api_response(success=False, message=str(e), code=500)
    finally:
        db.close()


# --- 批量更新/删除 ---

@app.route("/api/devices/batch", methods=["PUT"])
@login_required
def batch_update_devices():
    """批量更新设备字段：status / fault_disposition / operator"""
    data = request.get_json(force=True)
    ids = data.get("ids", [])
    field = data.get("field", "")
    value = data.get("value", None)
    if not ids or field not in ("status", "fault_disposition", "operator", "device_type"):
        return api_response(success=False, message="参数无效", code=400)
    db = next(get_db())
    try:
        updated = db.query(DeviceTest).filter(DeviceTest.id.in_(ids)).update(
            {field: value}, synchronize_session=False
        )
        db.commit()
        return api_response(data={"updated": updated}, message=f"已更新 {updated} 台设备")
    except Exception as e:
        db.rollback()
        return api_response(success=False, message=str(e), code=500)
    finally:
        db.close()


@app.route("/api/devices/batch", methods=["DELETE"])
@login_required
def batch_delete_devices():
    """批量删除设备"""
    data = request.get_json(force=True)
    ids = data.get("ids", [])
    if not ids:
        return api_response(success=False, message="请选择要删除的设备", code=400)
    db = next(get_db())
    try:
        # 记录 batch_id 用于更新计数
        batch_ids = db.query(DeviceTest.batch_id).filter(DeviceTest.id.in_(ids)).all()
        batch_id_set = set(bid[0] for bid in batch_ids if bid[0] is not None)

        # 先删除关联的测试结果和返厂记录
        db.query(DeviceTestResult).filter(DeviceTestResult.device_id.in_(ids)).delete(synchronize_session=False)
        db.query(DeviceReturnRecord).filter(DeviceReturnRecord.device_id.in_(ids)).delete(synchronize_session=False)

        deleted = db.query(DeviceTest).filter(DeviceTest.id.in_(ids)).delete(synchronize_session=False)
        # 更新批次 device_count
        for bid in batch_id_set:
            cnt = db.query(func.count(DeviceTest.id)).filter(DeviceTest.batch_id == bid).scalar() or 0
            db.query(Batch).filter(Batch.id == bid).update({"device_count": cnt})
        db.commit()
        return api_response(data={"deleted": deleted}, message=f"已删除 {deleted} 台设备")
    except Exception as e:
        db.rollback()
        return api_response(success=False, message=str(e), code=500)
    finally:
        db.close()


@app.route("/api/devices/batch/results", methods=["PUT"])
@login_required
def batch_update_results():
    """批量更新设备测试结果（支持单指标模式 + 多指标批量模式）"""
    data = request.get_json(force=True)
    ids = data.get("ids", [])
    # New multi-metric mode: results=[{standard_id, metric_name, pass}, ...]
    results = data.get("results", [])
    # Legacy single-metric mode
    standard_id = data.get("standard_id")
    metric_name = data.get("metric_name", "")
    pass_val = 1 if data.get("pass") else 0

    if not ids:
        return api_response(success=False, message="缺少设备 ID", code=400)

    # Normalize to results list
    if results and len(results) > 0:
        # Multi-metric mode
        work_items = []
        for r in results:
            work_items.append({
                "standard_id": r.get("standard_id"),
                "metric_name": r.get("metric_name", ""),
                "pass_": 1 if r.get("pass") else 0
            })
    elif standard_id and metric_name:
        # Legacy single-metric mode
        work_items = [{"standard_id": standard_id, "metric_name": metric_name, "pass_": pass_val}]
    else:
        return api_response(success=False, message="参数不完整", code=400)

    db = next(get_db())
    try:
        updated = 0
        for device_id in ids:
            for item in work_items:
                existing = db.query(DeviceTestResult).filter(
                    DeviceTestResult.device_id == device_id,
                    DeviceTestResult.standard_id == item["standard_id"],
                    DeviceTestResult.metric_name == item["metric_name"],
                ).first()
                if existing:
                    existing.pass_ = item["pass_"]
                else:
                    std = db.query(TestStandard).filter(TestStandard.id == item["standard_id"]).first()
                    expected = ""
                    if std and std.metrics:
                        metrics = json.loads(std.metrics)
                        for m in metrics:
                            if m.get("name") == item["metric_name"]:
                                expected = m.get("value", "")
                                break
                    r = DeviceTestResult(
                        device_id=device_id,
                        standard_id=item["standard_id"],
                        metric_name=item["metric_name"],
                        metric_value="",
                        expected_value=expected,
                        pass_=item["pass_"],
                        notes="",
                    )
                    db.add(r)
                updated += 1
        db.commit()
        return api_response(data={"updated": updated}, message=f"已更新 {updated} 条指标结果")
    except Exception as e:
        db.rollback()
        return api_response(success=False, message=str(e), code=500)
    finally:
        db.close()


# --- 统计接口 ---

@app.route("/api/statistics/overview", methods=["GET"])
@login_required
def statistics_overview():
    db = next(get_db())
    try:
        normal = db.query(func.count(DeviceTest.id)).filter(DeviceTest.status == "normal").scalar()
        fault = db.query(func.count(DeviceTest.id)).filter(DeviceTest.status == "fault").scalar()
        pending = db.query(func.count(DeviceTest.id)).filter(DeviceTest.status == "pending").scalar()
        total_devices = normal + fault + pending

        # 全部批次的板卡总数
        total_expected = db.query(func.sum(Batch.total_count)).scalar() or 0
        untested = max(0, total_expected - total_devices)

        # 互斥 disposition 统计（每台设备只归属一种 disposition，避免饼图重叠）
        disp_await = db.query(func.count(DeviceTest.id)).filter(DeviceTest.fault_disposition == "待返厂").scalar()
        disp_intransit = db.query(func.count(DeviceTest.id)).filter(DeviceTest.fault_disposition == "返厂中").scalar()
        disp_done = db.query(func.count(DeviceTest.id)).filter(DeviceTest.fault_disposition == "已返厂").scalar()
        disp_pending = db.query(func.count(DeviceTest.id)).filter(DeviceTest.fault_disposition == "pending").scalar()
        disp_stored = db.query(func.count(DeviceTest.id)).filter(DeviceTest.fault_disposition == "stored").scalar()

        # 故障率：待测+未测计入故障 (与批次列表计算规则一致)
        effective_total = max(total_expected, total_devices)
        fault_rate = float(round((effective_total - normal) / effective_total * 100, 2)) if effective_total > 0 else 0.0

        trend = []
        for i in range(29, -1, -1):
            d = _today_cn() - timedelta(days=i)
            day_total = db.query(func.count(DeviceTest.id)).filter(
                func.date(DeviceTest.test_date) == d
            ).scalar()
            day_fault = db.query(func.count(DeviceTest.id)).filter(
                func.date(DeviceTest.test_date) == d,
                DeviceTest.status == "fault",
            ).scalar()
            day_pending = db.query(func.count(DeviceTest.id)).filter(
                func.date(DeviceTest.test_date) == d,
                DeviceTest.status == "pending",
            ).scalar()
            trend.append({"date": d.isoformat(), "total": day_total, "fault": day_fault, "pending": day_pending})

        return api_response(
            data={
                "total": total_devices,
                "total_expected": total_expected,
                "normal": normal,
                "fault": fault,
                "pending": pending,
                "untested": untested,
                "dispositions": {
                    "待返厂": disp_await,
                    "返厂中": disp_intransit,
                    "已返厂": disp_done,
                    "pending": disp_pending,
                    "stored": disp_stored,
                },
                "fault_rate": fault_rate,
                "trend_30d": trend,
            }
        )
    finally:
        db.close()


@app.route("/api/statistics/daily", methods=["GET"])
@login_required
def daily_statistics():
    db = next(get_db())
    try:
        rows = (
            db.query(
                func.date(DeviceTest.test_date).label("day"),
                func.count(DeviceTest.id).label("total"),
                func.sum(case((DeviceTest.status == "normal", 1), else_=0)).label("normal"),
                func.sum(case((DeviceTest.status == "fault", 1), else_=0)).label("fault"),
                func.sum(case((DeviceTest.status == "pending", 1), else_=0)).label("status_pending"),
                func.sum(case((DeviceTest.fault_disposition.in_(["待返厂", "返厂中", "已返厂"]), 1), else_=0)).label("returned"),
                func.sum(case((DeviceTest.fault_disposition == "pending", 1), else_=0)).label("pending"),
                func.sum(case((DeviceTest.fault_disposition == "stored", 1), else_=0)).label("stored"),
            )
            .group_by(func.date(DeviceTest.test_date))
            .order_by(desc(func.date(DeviceTest.test_date)))
            .limit(90)
            .all()
        )
        result = []
        for r in rows:
            result.append({
                "date": str(r.day),
                "total": r.total,
                "normal": int(r.normal or 0),
                "fault": int(r.fault or 0),
                "status_pending": int(r.status_pending or 0),
                "returned": int(r.returned or 0),
                "pending": int(r.pending or 0),
                "stored": int(r.stored or 0),
            })
        return api_response(data=result)
    finally:
        db.close()


@app.route("/api/statistics/operators", methods=["GET"])
@login_required
def statistics_operators():
    db = next(get_db())
    try:
        rows = (
            db.query(
                DeviceTest.operator,
                func.count(DeviceTest.id).label("total"),
                func.sum(DeviceTest.status == "fault").label("fault"),
            )
            .filter(DeviceTest.operator.isnot(None), DeviceTest.operator != "")
            .group_by(DeviceTest.operator)
            .order_by(desc(func.count(DeviceTest.id)))
            .limit(10)
            .all()
        )
        result = [
            {"operator": r.operator, "total": r.total, "fault": int(r.fault or 0)}
            for r in rows
        ]
        return api_response(data=result)
    finally:
        db.close()


@app.route("/api/statistics/return-summary", methods=["GET"])
@login_required
def statistics_return_summary():
    db = next(get_db())
    try:
        total_returns = db.query(func.count(ReturnRecord.id)).scalar() or 0
        completed = db.query(func.count(ReturnRecord.id)).filter(ReturnRecord.status == "完成").scalar() or 0
        avg_days = 0.0
        if completed > 0:
            rows = (
                db.query(
                    func.datediff(ReturnRecord.updated_at, ReturnRecord.created_at)
                )
                .filter(ReturnRecord.status == "完成")
                .all()
            )
            if rows:
                avg_days = round(sum(r[0] for r in rows if r[0] is not None) / completed, 1)
        return api_response(
            data={
                "total": total_returns,
                "completed": completed,
                "avg_days": avg_days,
            }
        )
    finally:
        db.close()


@app.route("/api/statistics/heatmap", methods=["GET"])
@login_required
def statistics_heatmap():
    """返回操作热力图数据：按日-小时统计test_date的录入数量"""
    db = next(get_db())
    try:
        days = request.args.get("days", 30, type=int)
        rows = (
            db.query(
                func.date(DeviceTest.test_date).label("day"),
                func.hour(DeviceTest.test_date).label("hour"),
                func.count(DeviceTest.id).label("count"),
            )
            .filter(DeviceTest.test_date >= _today_cn() - timedelta(days=days))
            .group_by(func.date(DeviceTest.test_date), func.hour(DeviceTest.test_date))
            .order_by(func.date(DeviceTest.test_date), func.hour(DeviceTest.test_date))
            .all()
        )
        result = []
        for r in rows:
            result.append({
                "day": str(r.day),
                "hour": int(r.hour),
                "count": r.count,
            })
        return api_response(data=result)
    finally:
        db.close()


# --- 导出功能 ---

def _get_filtered_devices(db):
    query = db.query(DeviceTest)
    status_filter = request.args.get("status", "").strip()
    disposition_filter = request.args.get("disposition", "").strip()
    search = request.args.get("search", "").strip()
    date_from = request.args.get("date_from", "").strip()
    date_to = request.args.get("date_to", "").strip()

    if search:
        like = f"%{search}%"
        query = query.filter(
            or_(
                DeviceTest.board_mac.like(like),
                DeviceTest.wireless_mac.like(like),
                DeviceTest.ip_address.like(like),
                DeviceTest.return_tracking.like(like),
                DeviceTest.operator.like(like),
                DeviceTest.device_code.like(like),
            )
        )
    if status_filter:
        query = query.filter(DeviceTest.status == status_filter)
    if disposition_filter:
        if disposition_filter == "returned":
            disposition_filter = "待返厂"
        query = query.filter(DeviceTest.fault_disposition == disposition_filter)
    if date_from:
        query = query.filter(DeviceTest.test_date >= datetime.fromisoformat(date_from))
    if date_to:
        query = query.filter(DeviceTest.test_date <= datetime.fromisoformat(date_to + "T23:59:59"))

    return query.order_by(desc(DeviceTest.test_date)).all()


def _build_excel(devices, batch=None, db=None):
    """构建 Excel 工作簿。传入 batch 时生成双 Sheet 批次报告，否则生成单 Sheet 全局导出。"""
    wb = openpyxl.Workbook()

    if batch:
        # ---- Sheet1: 批次概要 ----
        ws1 = wb.active
        ws1.title = "批次概要"

        # Metadata
        total_devices = len(devices)
        normal_count = sum(1 for d in devices if d.status == "normal")
        fault_count = total_devices - normal_count
        progress_pct = round(normal_count / batch.total_count * 100, 1) if batch.total_count > 0 else 0
        est = batch.estimated_completion.isoformat() if batch.estimated_completion else ""
        act = batch.actual_completion.isoformat() if batch.actual_completion else ""
        overdue_note = ""
        if est and act and batch.actual_completion > batch.estimated_completion:
            overdue_note = batch.completion_note or f"实际完成 {act} 超过预计 {est}"

        header_font = Font(name="Helvetica Neue", bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_align = Alignment(vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin", color="B0C4DE"),
            right=Side(style="thin", color="B0C4DE"),
            top=Side(style="thin", color="B0C4DE"),
            bottom=Side(style="thin", color="B0C4DE"),
        )
        accent_fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")

        # Title row
        title_cell = ws1.cell(row=1, column=1, value=f"{batch.batch_no} 批次测试报告")
        title_cell.font = Font(name="Helvetica Neue", bold=True, size=14, color="1E3A5F")
        ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)

        info_rows = [
            ("批次号", batch.batch_no),
            ("批次描述", batch.description or ""),
            ("操作员", batch.operator),
            ("批次状态", batch.status),
            ("测试开始时间", batch.started_at.strftime("%Y-%m-%d %H:%M") if batch.started_at else ""),
            ("暂停/取消原因", batch.status_reason or ""),
            ("到货时间", batch.arrival_date.strftime("%Y-%m-%d %H:%M") if batch.arrival_date else ""),
            ("快递公司", batch.courier or ""),
            ("快递单号", batch.tracking_no or ""),
            ("测试标准", batch.test_standard or ""),
            ("测试时间", batch.test_date.strftime("%Y-%m-%d %H:%M") if batch.test_date else ""),
            ("板卡总数", batch.total_count or 0),
            ("已录入数", total_devices),
            ("正常设备", normal_count),
            ("故障设备", fault_count),
            ("测试进度", f"{progress_pct}% ({normal_count}/{batch.total_count or '?'})"),
            ("预计完成时间", est),
            ("实际完成时间", act),
            ("超期备注", overdue_note or batch.completion_note or ""),
        ]
        for i, (label, value) in enumerate(info_rows, 2):
            c1 = ws1.cell(row=i, column=1, value=label)
            c1.font = Font(name="Helvetica Neue", bold=True, size=10, color="333333")
            c1.fill = accent_fill
            c1.border = thin_border
            c1.alignment = Alignment(horizontal="right", vertical="center")
            c2 = ws1.cell(row=i, column=2, value=value)
            c2.font = Font(name="JetBrains Mono, Consolas, monospace", size=10)
            c2.border = thin_border
            c2.alignment = cell_align

        ws1.column_dimensions['A'].width = 16
        ws1.column_dimensions['B'].width = 45

        # ---- Sheet2: 设备明细 ----
        ws2 = wb.create_sheet("设备明细")
        headers2 = [
            "序号", "板卡MAC", "无线MAC", "IP地址", "状态",
            "故障原因/现象", "处置方式",
            "测试时间", "操作员", "备注",
        ]
        for col, h in enumerate(headers2, 1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        normal_fill = PatternFill(start_color="EAFAF1", end_color="EAFAF1", fill_type="solid")
        fault_fill = PatternFill(start_color="FDEDEC", end_color="FDEDEC", fill_type="solid")
        return_fill = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid")

        for i, d in enumerate(devices, 1):
            row_data = [
                i, d.board_mac, d.wireless_mac, d.ip_address,
                "正常" if d.status == "normal" else "故障",
                d.fault_reason or "",
                {"待返厂": "待返厂", "返厂中": "返厂中", "已返厂": "已返厂", "pending": "待处理", "stored": "已入库"}.get(d.fault_disposition, ""),
                d.test_date.strftime("%Y-%m-%d %H:%M") if d.test_date else "",
                d.operator or "",
                d.notes or "",
            ]
            row_fill = normal_fill if d.status == "normal" else (
                return_fill if d.fault_disposition in ("待返厂", "返厂中", "已返厂") else fault_fill
            )
            for col, val in enumerate(row_data, 1):
                cell = ws2.cell(row=i + 1, column=col, value=val)
                cell.alignment = cell_align
                cell.border = thin_border
                cell.fill = row_fill

        col_widths2 = [6, 18, 18, 16, 8, 30, 12, 18, 12, 20]
        for i, w in enumerate(col_widths2, 1):
            ws2.column_dimensions[get_column_letter(i)].width = w
        ws2.freeze_panes = "A2"
        ws2.auto_filter.ref = ws2.dimensions

        # ---- Sheet3: 测试标准报告 ----
        ws3 = wb.create_sheet("测试标准报告")
        headers3 = ["标准名称", "指标名称", "期望值", "通过设备数", "总设备数", "通过率"]
        for col, h in enumerate(headers3, 1):
            cell = ws3.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # 收集所有设备测试结果并按标准汇总
        device_ids = [d.id for d in devices]
        if device_ids:
            all_results = db.query(DeviceTestResult).filter(
                DeviceTestResult.device_id.in_(device_ids)
            ).all()
            # 按 (standard_id, metric_name) 汇总
            metric_stats = {}
            for r in all_results:
                key = (r.standard_id, r.metric_name, r.expected_value)
                if key not in metric_stats:
                    std_obj = db.query(TestStandard).filter(TestStandard.id == r.standard_id).first()
                    metric_stats[key] = {
                        "std_name": std_obj.name if std_obj else f"Standard-{r.standard_id}",
                        "total": 0,
                        "passed": 0,
                    }
                metric_stats[key]["total"] += 1
                if r.pass_:
                    metric_stats[key]["passed"] += 1

            row_idx = 2
            for (sid, mname, expected), stats in metric_stats.items():
                rate = round(stats["passed"] / stats["total"] * 100, 1) if stats["total"] > 0 else 0
                row_data = [stats["std_name"], mname, expected, stats["passed"], stats["total"], f"{rate}%"]
                for col, val in enumerate(row_data, 1):
                    cell = ws3.cell(row=row_idx, column=col, value=val)
                    cell.alignment = cell_align
                    cell.border = thin_border
                row_idx += 1

        col_widths3 = [18, 20, 20, 14, 14, 14]
        for i, w in enumerate(col_widths3, 1):
            ws3.column_dimensions[get_column_letter(i)].width = w
        ws3.freeze_panes = "A2"

    else:
        # Original global export
        ws = wb.active
        ws.title = "硬件测试记录"

        headers = [
            "序号", "设备编号", "板卡MAC", "无线MAC", "IP地址", "状态",
            "故障原因/现象", "处置方式",
            "测试日期", "操作员", "备注",
        ]

        header_font = Font(name="Helvetica Neue", bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_align = Alignment(vertical="center")
        thin_border = Border(
            left=Side(style="thin", color="D5D8DC"),
            right=Side(style="thin", color="D5D8DC"),
            top=Side(style="thin", color="D5D8DC"),
            bottom=Side(style="thin", color="D5D8DC"),
        )

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        normal_fill = PatternFill(start_color="EAFAF1", end_color="EAFAF1", fill_type="solid")
        fault_fill = PatternFill(start_color="FDEDEC", end_color="FDEDEC", fill_type="solid")
        return_fill = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid")

        for i, d in enumerate(devices, 1):
            row_data = [
                i, d.device_code or '', d.board_mac, d.wireless_mac, d.ip_address,
                "正常" if d.status == "normal" else "故障",
                d.fault_reason or "",
                {"待返厂": "待返厂", "返厂中": "返厂中", "已返厂": "已返厂", "pending": "待处理", "stored": "已入库"}.get(d.fault_disposition, ""),
                d.test_date.strftime("%Y-%m-%d %H:%M") if d.test_date else "",
                d.operator or "",
                d.notes or "",
            ]
            row_fill = normal_fill if d.status == "normal" else (
                return_fill if d.fault_disposition in ("待返厂", "返厂中", "已返厂") else fault_fill
            )
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=i + 1, column=col, value=val)
                cell.alignment = cell_align
                cell.border = thin_border
                cell.fill = row_fill

        col_widths = [6, 18, 18, 18, 16, 8, 30, 12, 18, 12, 20]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    return wb
def _build_markdown(devices):
    lines = [
        "# 硬件测试记录报告",
        "",
        f"> 导出时间: {_now_cn().strftime('%Y-%m-%d %H:%M:%S')}",
        f"> 记录总数: {len(devices)}",
        "",
        "## 统计概览",
        "",
    ]

    total = len(devices)
    normal = sum(1 for d in devices if d.status == "normal")
    fault = total - normal
    returned = sum(1 for d in devices if d.fault_disposition in ("待返厂", "返厂中", "已返厂"))
    pending = sum(1 for d in devices if d.fault_disposition == "pending")
    stored = sum(1 for d in devices if d.fault_disposition == "stored")

    lines.extend([
        f"| 指标 | 数量 | 占比 |",
        f"|------|------|------|",
        f"| 总记录 | {total} | 100% |",
        f"| 正常 | {normal} | {round(normal/total*100, 1) if total else 0}% |",
        f"| 故障 | {fault} | {round(fault/total*100, 1) if total else 0}% |",
        f"| 待返厂 | {returned} | - |",
        f"| 待处理 | {pending} | - |",
        f"| 已入库 | {stored} | - |",
        "",
        "## 详细记录",
        "",
    ])

    lines.append("| # | 板卡MAC | 无线MAC | IP地址 | 状态 | 故障原因 | 处置 |")
    lines.append("|---|---------|---------|--------|------|----------|------|")

    for i, d in enumerate(devices, 1):
        status_icon = "✓" if d.status == "normal" else "✗"
        disposition = {"待返厂": "待返厂", "返厂中": "返厂中", "已返厂": "已返厂", "pending": "待处理", "stored": "已入库"}.get(d.fault_disposition, "-")
        lines.append(
            f"| {i} | {d.board_mac} | {d.wireless_mac} | {d.ip_address} | {status_icon} | "
            f"{d.fault_reason or '-'} | {disposition} |"
        )

    return "\n".join(lines)


@app.route("/api/export/excel", methods=["GET"])
@login_required
def export_excel():
    db = next(get_db())
    try:
        devices = _get_filtered_devices(db)
        wb = _build_excel(devices, db=db)
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f"hardware_test_report_{_now_cn().strftime('%Y%m%d_%H%M%S')}.xlsx",
        )
    except Exception as e:
        return api_response(success=False, message=str(e), code=500)
    finally:
        db.close()



@app.route("/api/batches/<int:batch_id>/export/excel", methods=["GET"])
@login_required
def export_batch_excel(batch_id):
    """导出批次测试报告为 Excel（双 Sheet）。"""
    db = next(get_db())
    try:
        batch = db.query(Batch).filter(Batch.id == batch_id).first()
        if not batch:
            return api_response(success=False, message="批次不存在", code=404)
        devices = db.query(DeviceTest).filter(DeviceTest.batch_id == batch_id).order_by(desc(DeviceTest.test_date)).all()
        wb = _build_excel(devices, batch=batch, db=db)
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f"{batch.batch_no}_测试报告_{_today_cn().isoformat()}.xlsx",
        )
    finally:
        db.close()

@app.route("/api/export/excel-custom", methods=["POST"])
@login_required
def export_excel_custom():
    """自定义导出 Excel，支持列选择、表头重命名和去重。"""
    data = request.get_json() or {}
    devices_data = data.get("devices", [])
    filter_duplicates = data.get("filter_duplicates", False)
    selected_columns = data.get("columns", [])
    header_overrides = data.get("header_overrides", {})

    if not devices_data:
        return api_response(success=False, message="没有可导出的数据", code=400)

    if filter_duplicates:
        seen = set()
        unique_devices = []
        for d in devices_data:
            key = d.get("device_code", "") or d.get("board_mac", "")
            if key and key not in seen:
                seen.add(key)
                unique_devices.append(d)
        devices_data = unique_devices

    DEVICE_TYPE_MAP = {"1": "小周天", "2": "大周天", "3": "周天", "4": "智瞳"}
    column_config = {
        "device_code": ("设备编号", lambda d: d.get("device_code", "")),
        "device_type": ("设备类型", lambda d: d.get("device_type", "")),
        "board_mac": ("物理MAC", lambda d: d.get("board_mac", "")),
        "wireless_mac": ("MAC地址", lambda d: d.get("wireless_mac", "")),
        "ip_address": ("IP地址", lambda d: d.get("ip_address", "")),
        "batch_no": ("所属批次", lambda d: d.get("batch_no", "")),
        "status": ("状态", lambda d: {"normal": "正常", "fault": "故障", "pending": "待处理"}.get(d.get("status", ""), d.get("status", ""))),
        "disposal_type": ("处置方式", lambda d: d.get("disposal_type", "") or d.get("fault_disposition", "")),
        "fault_disposition": ("处置方式", lambda d: {"待返厂": "待返厂", "返厂中": "返厂中", "已返厂": "已返厂", "pending": "待处理", "stored": "已入库"}.get(d.get("fault_disposition", ""), "")),
        "operator": ("操作员", lambda d: d.get("operator", "") or d.get("tester", "")),
        "test_date": ("测试时间", lambda d: d.get("test_date", "")),
        "fault_reason": ("故障原因", lambda d: d.get("fault_reason", "")),
        "notes": ("备注", lambda d: d.get("notes", "") or d.get("remark", "")),
        "flow_status": ("流转状态", lambda d: d.get("flow_status", "")),
        "flow_purpose": ("流转用途", lambda d: d.get("flow_purpose", "")),
        "flow_destination": ("目的地", lambda d: d.get("flow_destination", "")),
        "flow_out_date": ("外出日期", lambda d: d.get("flow_out_date", "")),
        "flow_expected_return": ("预计返还", lambda d: d.get("flow_expected_return", "")),
    }

    if not selected_columns:
        selected_columns = ["board_mac", "wireless_mac", "ip_address", "batch_no", "status", "fault_disposition", "operator", "test_date", "fault_reason", "notes"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "设备列表"

    header_font = Font(name="Helvetica Neue", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_align = Alignment(vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="B0C4DE"),
        right=Side(style="thin", color="B0C4DE"),
        top=Side(style="thin", color="B0C4DE"),
        bottom=Side(style="thin", color="B0C4DE"),
    )

    for col_idx, col_key in enumerate(selected_columns, 1):
        default_header = column_config.get(col_key, (col_key, None))[0]
        header_text = header_overrides.get(col_key, default_header)
        cell = ws.cell(row=1, column=col_idx, value=header_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    normal_fill = PatternFill(start_color="EAFAF1", end_color="EAFAF1", fill_type="solid")
    fault_fill = PatternFill(start_color="FDEDEC", end_color="FDEDEC", fill_type="solid")

    for row_idx, device in enumerate(devices_data, 2):
        status_val = device.get("status", "")
        row_fill = normal_fill if status_val in ("normal", "正常") else fault_fill
        for col_idx, col_key in enumerate(selected_columns, 1):
            _, extractor = column_config.get(col_key, (col_key, lambda d: ""))
            value = extractor(device)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = cell_align
            cell.border = thin_border
            cell.fill = row_fill

    for col_idx, col_key in enumerate(selected_columns, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 16

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"设备列表_{_today_cn().isoformat()}.xlsx",
    )

@app.route("/api/export/markdown", methods=["GET"])
@login_required
def export_markdown():
    db = next(get_db())
    try:
        devices = _get_filtered_devices(db)
        md_content = _build_markdown(devices)
        output = io.BytesIO()
        output.write(md_content.encode("utf-8"))
        output.seek(0)
        return send_file(
            output,
            mimetype="text/markdown",
            as_attachment=True,
            download_name=f"hardware_test_report_{_now_cn().strftime('%Y%m%d_%H%M%S')}.md",
        )
    except Exception as e:
        return api_response(success=False, message=str(e), code=500)
    finally:
        db.close()


def _export_device_tests_to_excel(records, filename_prefix="exports"):
    """将 DeviceTest 记录导出为 Excel 并返回 Flask Response"""
    wb = _build_excel(records)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"{filename_prefix}_{_now_cn().strftime('%Y%m%d_%H%M%S')}.xlsx",
    )

def _export_return_records_to_excel(records, filename_prefix="return_records"):
    """将 ReturnRecord 记录导出为 Excel 并返回 Flask Response"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "返厂跟踪记录"

    headers = ["序号", "快递公司", "快递单号", "返厂状态", "关联设备数", "创建时间", "更新时间"]
    header_font = Font(name="Helvetica Neue", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_align = Alignment(vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="D5D8DC"),
        right=Side(style="thin", color="D5D8DC"),
        top=Side(style="thin", color="D5D8DC"),
        bottom=Side(style="thin", color="D5D8DC"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    status_fill = {
        "进行中": PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid"),
        "完成": PatternFill(start_color="EAFAF1", end_color="EAFAF1", fill_type="solid"),
        "取消": PatternFill(start_color="FDEDEC", end_color="FDEDEC", fill_type="solid"),
    }

    for i, r in enumerate(records, 1):
        row_data = [
            i,
            r.courier or "",
            r.tracking_no,
            r.status or "",
            r.device_count,
            r.created_at.strftime("%Y-%m-%d %H:%M") if r.created_at else "",
            r.updated_at.strftime("%Y-%m-%d %H:%M") if r.updated_at else "",
        ]
        row_fill = status_fill.get(r.status, PatternFill())
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i + 1, column=col, value=val)
            cell.alignment = cell_align
            cell.border = thin_border
            if row_fill:
                cell.fill = row_fill

    col_widths = [6, 16, 20, 12, 14, 20, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"{filename_prefix}_{_now_cn().strftime('%Y%m%d_%H%M%S')}.xlsx",
    )


# --- 设备测试记录批量导出 ---

@app.route("/api/device-tests/export-selected", methods=["POST"])
@login_required
def export_device_tests_selected():
    """导出选中的设备测试记录"""
    data = request.get_json(silent=True) or {}
    ids = data.get("ids", [])
    if not ids or not isinstance(ids, list):
        return api_response(message="请选择要导出的记录", code=400)
    db = next(get_db())
    try:
        records = db.query(DeviceTest).filter(DeviceTest.id.in_(ids)).order_by(DeviceTest.sort_order.is_(None), DeviceTest.sort_order.asc(), DeviceTest.id.asc()).all()
        return _export_device_tests_to_excel(records, filename_prefix="selected_tests")
    except Exception as e:
        return api_response(success=False, message=str(e), code=500)
    finally:
        db.close()

@app.route("/api/device-tests/export-batch", methods=["POST"])
@login_required
def export_device_tests_batch():
    """按筛选条件批量导出设备测试记录"""
    data = request.get_json(silent=True) or {}
    filters = data.get("filters", {})
    db = next(get_db())
    try:
        query = db.query(DeviceTest)
        if filters.get("status"):
            query = query.filter(DeviceTest.status == filters["status"])
        if filters.get("disposition"):
            disposition = filters["disposition"]
            if disposition == "returned":
                disposition = "待返厂"
            query = query.filter(DeviceTest.fault_disposition == disposition)
        if filters.get("keyword"):
            like = f"%{filters['keyword']}%"
            query = query.filter(or_(
                DeviceTest.board_mac.like(like),
                DeviceTest.wireless_mac.like(like),
                DeviceTest.ip_address.like(like),
                DeviceTest.operator.like(like),
                DeviceTest.device_code.like(like),
                DeviceTest.return_tracking.like(like),
            ))
        if filters.get("date_from"):
            query = query.filter(DeviceTest.test_date >= filters["date_from"])
        if filters.get("date_to"):
            query = query.filter(DeviceTest.test_date <= filters["date_to"] + " 23:59:59")
        records = query.order_by(DeviceTest.sort_order.is_(None), DeviceTest.sort_order.asc(), DeviceTest.id.asc()).all()
        return _export_device_tests_to_excel(records, filename_prefix="batch_tests")
    except Exception as e:
        return api_response(success=False, message=str(e), code=500)
    finally:
        db.close()

# --- 故障管理批量导出 ---

@app.route("/api/faults/export-selected", methods=["POST"])
@login_required
def export_faults_selected():
    """导出选中的故障记录"""
    data = request.get_json(silent=True) or {}
    ids = data.get("ids", [])
    if not ids or not isinstance(ids, list):
        return api_response(message="请选择要导出的记录", code=400)
    db = next(get_db())
    try:
        records = db.query(DeviceTest).filter(
            DeviceTest.id.in_(ids),
            DeviceTest.status == "fault"
        ).all()
        return _export_device_tests_to_excel(records, filename_prefix="selected_faults")
    except Exception as e:
        return api_response(success=False, message=str(e), code=500)
    finally:
        db.close()

@app.route("/api/faults/export-batch", methods=["POST"])
@login_required
def export_faults_batch():
    """按筛选条件批量导出故障记录"""
    data = request.get_json(silent=True) or {}
    filters = data.get("filters", {})
    db = next(get_db())
    try:
        query = db.query(DeviceTest).filter(DeviceTest.status == "fault")
        if filters.get("keyword"):
            like = f"%{filters['keyword']}%"
            query = query.filter(or_(
                DeviceTest.board_mac.like(like),
                DeviceTest.wireless_mac.like(like),
                DeviceTest.ip_address.like(like),
                DeviceTest.fault_reason.like(like),
                DeviceTest.return_tracking.like(like),
                DeviceTest.operator.like(like),
            ))
        if filters.get("disposition"):
            disposition = filters["disposition"]
            if disposition == "returned":
                disposition = "待返厂"
            query = query.filter(DeviceTest.fault_disposition == disposition)
        if filters.get("date_from"):
            query = query.filter(DeviceTest.test_date >= filters["date_from"])
        if filters.get("date_to"):
            query = query.filter(DeviceTest.test_date <= filters["date_to"] + " 23:59:59")
        records = query.order_by(desc(DeviceTest.test_date)).all()
        return _export_device_tests_to_excel(records, filename_prefix="batch_faults")
    except Exception as e:
        return api_response(success=False, message=str(e), code=500)
    finally:
        db.close()

# --- 返厂跟踪批量导出 ---

@app.route("/api/returns/export-selected", methods=["POST"])
@login_required
def export_returns_selected():
    """导出选中的返厂记录"""
    data = request.get_json(silent=True) or {}
    ids = data.get("ids", [])
    if not ids or not isinstance(ids, list):
        return api_response(message="请选择要导出的记录", code=400)
    db = next(get_db())
    try:
        records = db.query(ReturnRecord).filter(ReturnRecord.id.in_(ids)).all()
        return _export_return_records_to_excel(records, filename_prefix="selected_returns")
    except Exception as e:
        return api_response(success=False, message=str(e), code=500)
    finally:
        db.close()

@app.route("/api/returns/export-batch", methods=["POST"])
@login_required
def export_returns_batch():
    """按筛选条件批量导出返厂记录"""
    data = request.get_json(silent=True) or {}
    filters = data.get("filters", {})
    db = next(get_db())
    try:
        query = db.query(ReturnRecord)
        if filters.get("keyword"):
            like = f"%{filters['keyword']}%"
            if filters.get("search_field") == "courier":
                query = query.filter(ReturnRecord.courier.like(like))
            elif filters.get("search_field") == "tracking_no":
                query = query.filter(ReturnRecord.tracking_no.like(like))
            else:
                query = query.filter(or_(
                    ReturnRecord.tracking_no.like(like),
                    ReturnRecord.courier.like(like),
                ))
        records = query.order_by(desc(ReturnRecord.created_at)).all()
        return _export_return_records_to_excel(records, filename_prefix="batch_returns")
    except Exception as e:
        return api_response(success=False, message=str(e), code=500)
    finally:
        db.close()


# --- 自动完成接口 ---

@app.route("/api/suggest/operators", methods=["GET"])
@login_required
def suggest_operators():
    db = next(get_db())
    try:
        rows = (
            db.query(OperatorHistory)
            .order_by(desc(OperatorHistory.use_count), desc(OperatorHistory.last_used))
            .limit(20)
            .all()
        )
        return api_response(data=[r.to_dict() for r in rows])
    finally:
        db.close()


@app.route("/api/operators", methods=["POST"])
@login_required
def record_operator():
    data = request.get_json(force=True)
    name = data.get("name", "").strip()
    if not name:
        return api_response(success=False, message="操作员名称不能为空", code=400)

    db = next(get_db())
    try:
        operator = db.query(OperatorHistory).filter(OperatorHistory.name == name).first()
        if operator:
            operator.use_count = (operator.use_count or 0) + 1
            operator.last_used = _now_cn()
        else:
            operator = OperatorHistory(name=name, use_count=1, last_used=_now_cn())
            db.add(operator)
        db.commit()
        db.refresh(operator)
        return api_response(data=operator.to_dict(), message="操作员记录成功")
    except Exception as e:
        db.rollback()
        return api_response(success=False, message=str(e), code=500)
    finally:
        db.close()


# --- 模板下载 ---

@app.route("/api/template/download", methods=["GET"])
@login_required
def download_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "批量导入模板"

    headers = [
        "设备编号", "物理MAC", "无线MAC", "IP地址", "设备类型", "状态",
        "故障原因/现象", "处置方式",
        "测试时间", "操作员", "备注",
    ]

    header_font = Font(name="Helvetica Neue", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D5D8DC"),
        right=Side(style="thin", color="D5D8DC"),
        top=Side(style="thin", color="D5D8DC"),
        bottom=Side(style="thin", color="D5D8DC"),
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    example_row = [
        "留空自动生成", "00:1A:2B:3C:4D:5E", "00:1A:2B:3C:4D:5F", "192.168.1.100", "1=小周天/2=大周天/3=周天/4=智瞳", "正常",
        "", "",
        _now_cn().strftime("%Y-%m-%d %H:%M:%S"), "操作员姓名", "",
    ]
    for col, val in enumerate(example_row, 1):
        cell = ws.cell(row=2, column=col, value=val)
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center")

    col_widths = [14, 18, 18, 16, 24, 8, 30, 12, 20, 14, 18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"batch_import_template_{_now_cn().strftime('%Y%m%d')}.xlsx",
    )


# --- 故障报备 API ---

@app.route("/api/devices/<int:device_id>/fault", methods=["PUT"])
@login_required
def report_fault(device_id):
    db = next(get_db())
    try:
        device = db.query(DeviceTest).filter(DeviceTest.id == device_id).first()
        if not device:
            return api_response(success=False, message="设备记录不存在", code=404)

        data = request.get_json(force=True)

        device.status = "fault"
        device.fault_reason = (data.get("fault_reason") or "").strip() or None
        device.fault_disposition = data.get("fault_disposition") or 'pending'

        return_date = data.get("return_date")
        device.return_date = datetime.strptime(return_date, "%Y-%m-%d").date() if return_date else None
        device.return_courier = (data.get("return_courier") or "").strip() or None
        device.return_tracking = (data.get("return_tracking") or "").strip() or None

        db.commit()
        db.refresh(device)
        return api_response(data=device.to_dict(), message="故障报备成功")
    except Exception as e:
        db.rollback()
        return api_response(success=False, message=str(e), code=500)
    finally:
        db.close()


# --- 故障设备列表 API ---

@app.route("/api/faults", methods=["GET"])
@login_required
def list_faults():
    db = next(get_db())
    try:
        page = request.args.get("page", 1, type=int)
        per_page = request.args.get("per_page", 20, type=int)
        search = request.args.get("search", "").strip()
        disposition = request.args.get("disposition", "").strip()
        sort_by = request.args.get("sort_by", "test_date")
        sort_order = request.args.get("sort_order", "desc")

        query = db.query(DeviceTest).filter(DeviceTest.status == "fault")

        # 搜索过滤
        if search:
            query = query.filter(
                or_(
                    DeviceTest.board_mac.contains(search),
                    DeviceTest.wireless_mac.contains(search),
                    DeviceTest.ip_address.contains(search),
                    DeviceTest.fault_reason.contains(search),
                    DeviceTest.return_tracking.contains(search),
                    DeviceTest.operator.contains(search),
                )
            )

        # 处置状态过滤（前端可能传旧值 "returned"，映射为 "待返厂"）
        if disposition:
            if disposition == "returned":
                disposition = "待返厂"
            query = query.filter(DeviceTest.fault_disposition == disposition)

        # 排序
        sort_col = getattr(DeviceTest, sort_by, DeviceTest.test_date)
        if sort_order == "asc":
            query = query.order_by(sort_col.asc())
        else:
            query = query.order_by(sort_col.desc())

        total = query.count()
        items = query.offset((page - 1) * per_page).limit(per_page).all()

        # 为每个设备附加批次号
        device_list = []
        for d in items:
            d_dict = d.to_dict()
            d_dict["batch_no"] = d.batch.batch_no if d.batch else ""
            device_list.append(d_dict)

        return api_response(data={
            "items": device_list,
            "total": total,
            "page": page,
            "per_page": per_page,
        })
    except Exception as e:
        return api_response(success=False, message=str(e), code=500)
    finally:
        db.close()


# ===========================================================================
# 系统配置 API
# ===========================================================================

@app.route("/api/config", methods=["GET"])
@admin_required
def get_configs():
    """返回所有系统配置的 {key: value} 字典"""
    db = next(get_db())
    try:
        configs = db.query(SystemConfig).all()
        result = {c.config_key: c.config_value for c in configs}
        return api_response(data=result)
    finally:
        db.close()


@app.route("/api/config", methods=["PUT"])
@admin_required
def update_configs():
    """批量 upsert 配置项，接收 {key: value} 字典"""
    data = request.get_json(force=True) or {}
    if not isinstance(data, dict) or not data:
        return api_response(success=False, message="请提供有效的配置字典", code=400)

    db = next(get_db())
    try:
        for key, value in data.items():
            key = str(key).strip()
            if not key:
                continue
            value = str(value) if value is not None else ""
            existing = db.query(SystemConfig).filter(SystemConfig.config_key == key).first()
            if existing:
                existing.config_value = value
                existing.updated_at = _now_cn()
            else:
                db.add(SystemConfig(config_key=key, config_value=value))
        db.commit()
        return api_response(message="配置已保存")
    except Exception as e:
        db.rollback()
        return api_response(success=False, message=str(e), code=500)
    finally:
        db.close()


# ===========================================================================
# 物流查询 API
# ===========================================================================

import urllib.request
import urllib.parse
import hashlib


@app.route("/api/logistics/query", methods=["GET"])
@login_required
def query_logistics():
    """
    快递物流查询（快递100企业版 API）:
    - 如果传了 tracking_no，先自动识别快递公司，再查询物流轨迹
    - 需要 SystemConfig 中配置 kuaidi100_customer 和 kuaidi100_key
    - 参考: https://api.kuaidi100.com/document/
    """
    tracking_no = request.args.get("tracking_no", "").strip()
    if not tracking_no:
        return api_response(success=False, message="请提供快递单号", code=400)

    db = next(get_db())
    try:
        configs = {c.config_key: c.config_value for c in db.query(SystemConfig).all()}
    finally:
        db.close()

    customer = configs.get("kuaidi100_customer", "").strip()
    api_key = configs.get("kuaidi100_key", "").strip()

    if not customer or not api_key:
        return api_response(success=False, message="请先在系统配置中设置快递100 API密钥 (customer + key)", code=400)

    def _make_sign(param_str, key, cust):
        """快递100企业版签名: MD5(param + key + customer) 后转大写"""
        raw = param_str + key + cust
        return hashlib.md5(raw.encode("utf-8")).hexdigest().upper()

    # ====== 智能单号识别（四步降级方案） ======
    com_code = None
    com_name = None

    # Step 1: 智能识别（优先尝试企业版 autonumber API）
    auto_url = "https://poll.kuaidi100.com/poll/autonumber.do"
    auto_param_str = json.dumps({"num": tracking_no}, separators=(",", ":"))
    auto_sign = _make_sign(auto_param_str, api_key, customer)
    auto_params = {
        "customer": customer,
        "param": auto_param_str,
        "sign": auto_sign,
    }
    try:
        auto_data = urllib.parse.urlencode(auto_params).encode("utf-8")
        auto_req = urllib.request.Request(auto_url, data=auto_data, method="POST")
        auto_req.add_header("Content-Type", "application/x-www-form-urlencoded")
        auto_req.add_header("User-Agent", "Mozilla/5.0 (compatible; HardwareTestSystem/1.0)")
        with urllib.request.urlopen(auto_req, timeout=10) as resp:
            result = json.loads(resp.read().decode("utf-8"))
            if result.get("auto") and isinstance(result["auto"], list) and len(result["auto"]) > 0:
                com_code = result["auto"][0].get("comCode", "")
                com_name = result["auto"][0].get("name", "")
            if not com_code and result.get("comCode"):
                com_code = result.get("comCode")
                com_name = result.get("name", "")
            if not com_code and isinstance(result, list) and len(result) > 0:
                com_code = result[0].get("comCode", "")
                com_name = result[0].get("name", "")
    except Exception:
        pass

    # Step 2: 前缀推断（降级方案A）
    if not com_code:
        prefix_map = {
            "SF": ("shunfeng", "顺丰速运"),
            "YT": ("yuantong", "圆通速递"),
            "ZT": ("zhongtong", "中通快递"),
            "JT": ("jtexpress", "极兔速递"),
            "STO": ("shentong", "申通快递"),
            "YD": ("yunda", "韵达快递"),
            "JD": ("jd", "京东物流"),
            "DB": ("debang", "德邦快递"),
            "EMS": ("ems", "EMS"),
            "99": ("youzhengguonei", "邮政国内"),
        }
        upper_no = tracking_no.upper()
        for prefix, (code, name) in prefix_map.items():
            if upper_no.startswith(prefix):
                com_code = code
                com_name = name
                break

    # Step 3: 逐个试探（降级方案B）
    if not com_code:
        trial_codes = [
            ("yuantong", "圆通速递"),
            ("zhongtong", "中通快递"),
            ("shentong", "申通快递"),
            ("yunda", "韵达快递"),
            ("shunfeng", "顺丰速运"),
            ("jtexpress", "极兔速递"),
            ("jd", "京东物流"),
            ("debang", "德邦快递"),
            ("ems", "EMS"),
            ("youzhengguonei", "邮政国内"),
        ]
        for trial_code, trial_name in trial_codes:
            try:
                trial_param_str = json.dumps({"com": trial_code, "num": tracking_no}, separators=(",", ":"))
                trial_sign = _make_sign(trial_param_str, api_key, customer)
                trial_params = {
                    "customer": customer,
                    "param": trial_param_str,
                    "sign": trial_sign,
                }
                trial_data = urllib.parse.urlencode(trial_params).encode("utf-8")
                trial_req = urllib.request.Request(
                    "https://poll.kuaidi100.com/poll/query.do", data=trial_data, method="POST"
                )
                trial_req.add_header("Content-Type", "application/x-www-form-urlencoded")
                trial_req.add_header("User-Agent", "Mozilla/5.0 (compatible; HardwareTestSystem/1.0)")
                with urllib.request.urlopen(trial_req, timeout=10) as resp:
                    trial_result = json.loads(resp.read().decode("utf-8"))
                    state = trial_result.get("state", "")
                    if trial_result.get("data") or state in ("0", "1", "2", "3", "4"):
                        com_code = trial_code
                        com_name = trial_name
                        break
            except Exception:
                continue

    # Step 4: 全部失败
    if not com_code:
        return api_response(
            success=False,
            message=f"无法识别快递单号 {tracking_no} 对应的快递公司（请确认单号正确，或检查系统设置中的快递100密钥）",
            code=400,
        )

    # Step 2: 查询物流轨迹（需 sign）
    query_url = "https://poll.kuaidi100.com/poll/query.do"
    query_param_str = json.dumps({"com": com_code, "num": tracking_no}, separators=(",", ":"))
    query_sign = _make_sign(query_param_str, api_key, customer)
    query_params = {
        "customer": customer,
        "param": query_param_str,
        "sign": query_sign,
    }

    try:
        query_data = urllib.parse.urlencode(query_params).encode("utf-8")
        query_req = urllib.request.Request(query_url, data=query_data, method="POST")
        query_req.add_header("Content-Type", "application/x-www-form-urlencoded")
        query_req.add_header("User-Agent", "Mozilla/5.0 (compatible; HardwareTestSystem/1.0)")
        with urllib.request.urlopen(query_req, timeout=10) as resp:
            logistics_data = json.loads(resp.read().decode("utf-8"))
    except Exception as e:
        return api_response(success=False, message=f"物流轨迹查询失败: {str(e)}", code=500)

    # 解析轨迹
    tracks = []
    raw_data = logistics_data.get("data")
    if raw_data and isinstance(raw_data, list):
        for item in raw_data:
            tracks.append({
                "time": item.get("time", ""),
                "status": item.get("status", ""),
                "context": item.get("context", ""),
                "location": item.get("location", ""),
            })

    return api_response(data={
        "com_code": com_code,
        "com_name": com_name or com_code,
        "tracking_no": tracking_no,
        "state": logistics_data.get("state", ""),
        "tracks": tracks,
    })


# ===========================================================================
# 返厂跟踪 API
# ===========================================================================

@app.route("/api/return-records", methods=["GET"])
@login_required
def list_return_records():
    """返厂记录列表（分页 + 搜索）"""
    db = next(get_db())
    try:
        page = request.args.get("page", 1, type=int)
        per_page = request.args.get("per_page", 20, type=int)
        search = request.args.get("search", "").strip()
        search_field = request.args.get("search_field", "all").strip()

        query = db.query(ReturnRecord)

        if search:
            if search_field == "courier":
                query = query.filter(ReturnRecord.courier.contains(search))
            elif search_field == "tracking_no":
                query = query.filter(ReturnRecord.tracking_no.contains(search))
            else:
                query = query.filter(
                    or_(
                        ReturnRecord.tracking_no.contains(search),
                        ReturnRecord.courier.contains(search),
                    )
                )

        query = query.order_by(ReturnRecord.created_at.desc())
        total = query.count()
        items = query.offset((page - 1) * per_page).limit(per_page).all()

        return api_response(data={
            "items": [r.to_dict() for r in items],
            "total": total,
            "page": page,
            "per_page": per_page,
        })
    except Exception as e:
        return api_response(success=False, message=str(e), code=500)
    finally:
        db.close()


@app.route("/api/return-records", methods=["POST"])
@login_required
def create_return_record():
    """新建返厂记录"""
    db = next(get_db())
    try:
        data = request.get_json(force=True)
        tracking_no = (data.get("tracking_no") or "").strip()
        courier = (data.get("courier") or "").strip()
        device_ids = data.get("device_ids") or []
        return_code = (data.get("return_code") or "").strip()

        if not tracking_no:
            return api_response(success=False, message="快递单号不能为空", code=400)

        # 检查单号唯一性
        existing = db.query(ReturnRecord).filter(ReturnRecord.tracking_no == tracking_no).first()
        if existing:
            return api_response(success=False, message=f"快递单号 {tracking_no} 已存在", code=400)

        # 自动生成返厂单编号（若前端未传入）
        if not return_code:
            prefix_config = db.query(SystemConfig).filter(
                SystemConfig.config_key == "return_code_prefix"
            ).first()
            prefix = prefix_config.config_value.strip() if prefix_config and prefix_config.config_value.strip() else "RT"
            today_str = _now_cn().strftime("%Y%m%d")
            code_prefix = f"{prefix}{today_str}"
            last_record = db.query(ReturnRecord).filter(
                ReturnRecord.return_code.like(f"{code_prefix}%")
            ).order_by(desc(ReturnRecord.return_code)).first()
            if last_record and last_record.return_code.startswith(code_prefix):
                try:
                    last_seq = int(last_record.return_code[-3:])
                    seq = last_seq + 1
                except ValueError:
                    seq = 1
            else:
                seq = 1
            return_code = f"{code_prefix}{seq:03d}"

        # 校验设备
        if not device_ids:
            return api_response(success=False, message="请至少选择一个设备", code=400)

        devices = db.query(DeviceTest).filter(
            DeviceTest.id.in_(device_ids),
            DeviceTest.status == "fault",
            DeviceTest.fault_disposition == "待返厂",
        ).all()

        if len(devices) != len(device_ids):
            return api_response(success=False, message="部分设备不符合关联条件（需状态为故障且处置为待返厂）", code=400)

        # 检查设备是否已被其他返厂记录关联
        already_linked = db.query(DeviceReturnRecord).filter(
            DeviceReturnRecord.device_id.in_(device_ids)
        ).all()
        if already_linked:
            linked_ids = [dr.device_id for dr in already_linked]
            return api_response(
                success=False,
                message=f"以下设备已关联其他返厂记录: {linked_ids}",
                code=400,
            )

        # 创建返厂记录
        record = ReturnRecord(
            return_code=return_code,
            tracking_no=tracking_no,
            courier=courier,
            status=data.get("status", "进行中"),
            device_count=len(devices),
        )
        db.add(record)
        db.flush()

        # 创建设备关联并更新设备字段
        for device in devices:
            dr = DeviceReturnRecord(device_id=device.id, return_record_id=record.id)
            db.add(dr)
            device.fault_return_courier = courier
            device.fault_return_tracking = return_code
            device.fault_disposition = '返厂中'

        db.commit()
        db.refresh(record)

        return api_response(data=record.to_detail_dict(), message="返厂记录创建成功", code=201)
    except Exception as e:
        db.rollback()
        return api_response(success=False, message=str(e), code=500)
    finally:
        db.close()


@app.route("/api/return-records/<int:record_id>", methods=["GET"])
@login_required
def get_return_record(record_id):
    """获取返厂记录详情"""
    db = next(get_db())
    try:
        record = db.query(ReturnRecord).filter(ReturnRecord.id == record_id).first()
        if not record:
            return api_response(success=False, message="返厂记录不存在", code=404)
        return api_response(data=record.to_detail_dict())
    finally:
        db.close()


@app.route("/api/return-records/<int:record_id>", methods=["PUT"])
@login_required
def update_return_record(record_id):
    """编辑返厂记录"""
    db = next(get_db())
    try:
        record = db.query(ReturnRecord).filter(ReturnRecord.id == record_id).first()
        if not record:
            return api_response(success=False, message="返厂记录不存在", code=404)

        data = request.get_json(force=True)
        new_return_code = (data.get("return_code") or "").strip()
        new_tracking_no = (data.get("tracking_no") or "").strip()
        new_courier = (data.get("courier") or "").strip()
        new_status = data.get("status", record.status)
        new_device_ids = data.get("device_ids") or []

        # 检查单号唯一性
        if new_tracking_no and new_tracking_no != record.tracking_no:
            existing = db.query(ReturnRecord).filter(
                ReturnRecord.tracking_no == new_tracking_no,
                ReturnRecord.id != record_id,
            ).first()
            if existing:
                return api_response(success=False, message=f"快递单号 {new_tracking_no} 已存在", code=400)

        # 更新基本信息
        if new_tracking_no:
            record.tracking_no = new_tracking_no
        if new_return_code:
            record.return_code = new_return_code
        record.courier = new_courier
        record.status = new_status

        # 如果返厂状态更新为"完成"，批量更新关联设备为"已返厂"
        if new_status == "完成":
            linked = db.query(DeviceReturnRecord).filter(
                DeviceReturnRecord.return_record_id == record_id
            ).all()
            for lr in linked:
                linked_dev = db.query(DeviceTest).filter(DeviceTest.id == lr.device_id).first()
                if linked_dev:
                    linked_dev.fault_disposition = '已返厂'

        # 处理设备关联变更
        old_links = db.query(DeviceReturnRecord).filter(
            DeviceReturnRecord.return_record_id == record_id
        ).all()
        old_device_ids = {dr.device_id for dr in old_links}

        new_device_id_set = set(new_device_ids)
        to_remove = old_device_ids - new_device_id_set
        to_add = new_device_id_set - old_device_ids

        # 校验新增设备
        if to_add:
            devices = db.query(DeviceTest).filter(
                DeviceTest.id.in_(list(to_add)),
                DeviceTest.status == "fault",
                DeviceTest.fault_disposition == "待返厂",
            ).all()
            if len(devices) != len(to_add):
                return api_response(success=False, message="部分设备不符合关联条件", code=400)

            already_linked = db.query(DeviceReturnRecord).filter(
                DeviceReturnRecord.device_id.in_(list(to_add)),
                DeviceReturnRecord.return_record_id != record_id,
            ).all()
            if already_linked:
                linked_ids = [dr.device_id for dr in already_linked]
                return api_response(
                    success=False,
                    message=f"以下设备已关联其他返厂记录: {linked_ids}",
                    code=400,
                )

        # 取消关联的设备：清空 fault_return_* 字段，恢复"待返厂"状态
        for device_id in to_remove:
            device = db.query(DeviceTest).filter(DeviceTest.id == device_id).first()
            if device:
                device.fault_return_courier = ""
                device.fault_return_tracking = ""
                device.fault_disposition = '待返厂'
            db.query(DeviceReturnRecord).filter(
                DeviceReturnRecord.device_id == device_id,
                DeviceReturnRecord.return_record_id == record_id,
            ).delete()

        # 新增关联设备
        for device_id in to_add:
            dr = DeviceReturnRecord(device_id=device_id, return_record_id=record_id)
            db.add(dr)
            device = db.query(DeviceTest).filter(DeviceTest.id == device_id).first()
            if device:
                device.fault_return_courier = record.courier
                device.fault_return_tracking = record.return_code
                device.fault_disposition = '返厂中'

        # 更新现有关联设备的快递信息
        for device_id in (old_device_ids & new_device_id_set):
            device = db.query(DeviceTest).filter(DeviceTest.id == device_id).first()
            if device:
                device.fault_return_courier = record.courier
                device.fault_return_tracking = record.return_code
                # 返厂单完成 → 设备已在前置逻辑置为"已返厂"，不再覆盖
                # 否则确保设备状态为"返厂中"
                if record.status != "完成" and device.fault_disposition != '返厂中':
                    device.fault_disposition = '返厂中'

        record.device_count = len(new_device_id_set)

        db.commit()
        db.refresh(record)
        return api_response(data=record.to_detail_dict(), message="返厂记录更新成功")
    except Exception as e:
        db.rollback()
        return api_response(success=False, message=str(e), code=500)
    finally:
        db.close()


@app.route("/api/return-records/<int:record_id>", methods=["DELETE"])
@login_required
def delete_return_record(record_id):
    """删除返厂记录"""
    db = next(get_db())
    try:
        record = db.query(ReturnRecord).filter(ReturnRecord.id == record_id).first()
        if not record:
            return api_response(success=False, message="返厂记录不存在", code=404)

        # 清空关联设备的 fault_return_* 字段，并恢复处置状态为"待返厂"
        links = db.query(DeviceReturnRecord).filter(
            DeviceReturnRecord.return_record_id == record_id
        ).all()
        for link in links:
            device = db.query(DeviceTest).filter(DeviceTest.id == link.device_id).first()
            if device:
                device.fault_return_courier = ""
                device.fault_return_tracking = ""
                device.fault_disposition = '待返厂'

        # 删除关联记录
        db.query(DeviceReturnRecord).filter(
            DeviceReturnRecord.return_record_id == record_id
        ).delete()

        # 删除返厂记录
        db.delete(record)
        db.commit()
        return api_response(message="返厂记录已删除")
    except Exception as e:
        db.rollback()
        return api_response(success=False, message=str(e), code=500)
    finally:
        db.close()


@app.route("/api/return-records/available-devices", methods=["GET"])
@login_required
def get_available_devices():
    """获取可关联设备列表（status='fault' 且 fault_disposition='待返厂' 且未被关联）"""
    db = next(get_db())
    try:
        exclude_record_id = request.args.get("exclude_record_id", type=int)

        # 获取已关联的设备ID，排除当前记录（编辑时）
        if exclude_record_id:
            linked_ids = [
                dr.device_id for dr in db.query(DeviceReturnRecord.device_id).filter(
                    DeviceReturnRecord.return_record_id != exclude_record_id
                ).all()
            ]
        else:
            linked_ids = [dr.device_id for dr in db.query(DeviceReturnRecord.device_id).all()]

        query = db.query(DeviceTest).filter(
            DeviceTest.status == "fault",
            DeviceTest.fault_disposition == "待返厂",
        )
        if linked_ids:
            query = query.filter(~DeviceTest.id.in_(linked_ids))

        devices = query.order_by(DeviceTest.id.desc()).all()

        return api_response(data=[d.to_dict() for d in devices])
    except Exception as e:
        return api_response(success=False, message=str(e), code=500)
    finally:
        db.close()


# ---------------------------------------------------------------------------
# 启动
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    # 创建/更新数据库表
    Base.metadata.create_all(bind=engine)

    # 迁移：为 device_tests 表添加 return_courier 字段（如果不存在）
    with engine.connect() as conn:
        try:
            conn.execute(text(
                "ALTER TABLE device_tests ADD COLUMN return_courier VARCHAR(64) DEFAULT '' COMMENT '返厂快递公司'"
            ))
            conn.commit()
            print("[INFO] device_tests.return_courier 字段已添加")
        except Exception as e:
            pass  # 字段已存在则跳过

    # 迁移：创建 system_configs 表（如果不存在）
    try:
        SystemConfig.__table__.create(bind=engine, checkfirst=True)
        print("[INFO] system_configs 表已创建（如不存在）")
    except Exception as e:
        print(f"[WARN] system_configs 表创建失败: {e}")

    # 迁移：插入默认时区配置（如果不存在）
    try:
        _db = SessionLocal()
        if not _db.query(SystemConfig).filter(SystemConfig.config_key == "timezone").first():
            _db.add(SystemConfig(config_key="timezone", config_value="Asia/Shanghai"))
            _db.commit()
            print("[INFO] 已插入默认时区配置: Asia/Shanghai")
        _db.close()
    except Exception as e:
        print(f"[WARN] 默认时区配置插入失败: {e}")

    # 迁移：为 batches 表添加新字段（如果不存在）
    with engine.connect() as conn:
        try:
            conn.execute(text(
                "ALTER TABLE batches ADD COLUMN estimated_completion DATE DEFAULT NULL COMMENT '测试预计完成时间'"
            ))
            conn.execute(text(
                "ALTER TABLE batches ADD COLUMN actual_completion DATE DEFAULT NULL COMMENT '实际完成时间'"
            ))
            conn.execute(text(
                "ALTER TABLE batches ADD COLUMN completion_note TEXT DEFAULT NULL COMMENT '超期备注说明'"
            ))
            conn.commit()
            print("[INFO] batches 表新增字段迁移完成")
        except Exception:
            pass  # 字段已存在则跳过

    # 迁移：批次状态枚举扩展 + 新增字段
    with engine.connect() as conn:
        try:
            # 尝试添加新列（MySQL 8 支持 ALTER TABLE ADD COLUMN IF NOT EXISTS）
            conn.execute(text(
                "ALTER TABLE batches ADD COLUMN started_at DATETIME DEFAULT NULL COMMENT '测试开始时间'"
            ))
            print("[INFO] batches.started_at 字段已添加")
        except Exception:
            pass
        try:
            conn.execute(text(
                "ALTER TABLE batches ADD COLUMN status_reason TEXT DEFAULT NULL COMMENT '暂停/取消原因'"
            ))
            print("[INFO] batches.status_reason 字段已添加")
        except Exception:
            pass
        try:
            # 先将 active 映射为 in_progress（必须在 ALTER 之前，否则旧值无法通过新枚举）
            conn.execute(text(
                "UPDATE batches SET status = 'in_progress' WHERE status = 'active'"
            ))
            # 修改 ENUM 类型
            conn.execute(text(
                "ALTER TABLE batches MODIFY COLUMN status "
                "ENUM('planned','in_progress','paused','completed','cancelled') "
                "DEFAULT 'planned' COMMENT '批次状态'"
            ))
            conn.commit()
            print("[INFO] batches.status 枚举已迁移（active → in_progress）")
        except Exception as e:
            print(f"[WARN] batches.status 枚举迁移失败（可能已完成）: {e}")
            conn.rollback()

    # 迁移：device_tests 表 status 枚举扩展
    with engine.connect() as conn:
        try:
            conn.execute(text(
                "ALTER TABLE device_tests MODIFY COLUMN status "
                "ENUM('normal','fault','pending') "
                "DEFAULT 'pending' COMMENT '设备状态'"
            ))
            conn.commit()
            print("[INFO] device_tests.status 枚举已迁移")
        except Exception as e:
            print(f"[WARN] device_tests.status 枚举迁移失败（可能已完成）: {e}")
            conn.rollback()

    # 创建 device_test_results 表（如果不存在）
    try:
        DeviceTestResult.__table__.create(bind=engine, checkfirst=True)
        print("[INFO] device_test_results 表已创建（如不存在）")
    except Exception as e:
        print(f"[WARN] device_test_results 表创建失败: {e}")

    # 迁移：创建 return_records 和 device_return_records 表
    try:
        ReturnRecord.__table__.create(bind=engine, checkfirst=True)
        print("[INFO] return_records 表已创建（如不存在）")
    except Exception as e:
        print(f"[WARN] return_records 表创建失败: {e}")
    try:
        DeviceReturnRecord.__table__.create(bind=engine, checkfirst=True)
        print("[INFO] device_return_records 表已创建（如不存在）")
    except Exception as e:
        print(f"[WARN] device_return_records 表创建失败: {e}")

    # 迁移：创建 user_command_pins 表（如果不存在）
    try:
        UserCommandPin.__table__.create(bind=engine, checkfirst=True)
        print("[INFO] user_command_pins 表已创建（如不存在）")
    except Exception as e:
        print(f"[WARN] user_command_pins 表创建失败: {e}")

    # 迁移：为 device_tests 添加 fault_return_courier / fault_return_tracking 字段
    with engine.connect() as conn:
        try:
            conn.execute(text(
                "ALTER TABLE device_tests ADD COLUMN fault_return_courier VARCHAR(50) DEFAULT '' COMMENT '返厂跟踪-快递公司'"
            ))
            conn.commit()
            print("[INFO] device_tests.fault_return_courier 字段已添加")
        except Exception:
            pass
        try:
            conn.execute(text(
                "ALTER TABLE device_tests ADD COLUMN fault_return_tracking VARCHAR(50) DEFAULT '' COMMENT '返厂跟踪-快递单号'"
            ))
            conn.commit()
            print("[INFO] device_tests.fault_return_tracking 字段已添加")
        except Exception:
            pass

    # 迁移：为 device_tests 添加 sort_order 字段
    with engine.connect() as conn:
        try:
            conn.execute(text(
                "ALTER TABLE device_tests ADD COLUMN sort_order INTEGER DEFAULT NULL COMMENT '排序序号'"
            ))
            conn.commit()
            print("[INFO] device_tests.sort_order 字段已添加")
        except Exception:
            pass

    # 迁移：为 device_tests 添加 device_code 字段
    with engine.connect() as conn:
        try:
            conn.execute(text(
                "ALTER TABLE device_tests ADD COLUMN device_code VARCHAR(64) DEFAULT '' COMMENT '设备编号（MAC去冒号大写）'"
            ))
            conn.commit()
            print("[INFO] device_tests.device_code 字段已添加")
        except Exception:
            pass

    # 注意: 默认管理员不再每次启动自动创建，通过 Web 界面管理
    # 首次部署请使用: python3 -c "..." 手动创建

    # 预置默认测试标准
    db = SessionLocal()
    try:
        existing_count = db.query(TestStandard).count()
        if existing_count == 0:
            defaults = [
                TestStandard(
                    name="良品率测试",
                    description="基于 IPC-A-610 / IPC-9252 标准的板卡生产良率评估，评估制造工艺质量和一次通过率。",
                    metrics=json.dumps([
                        {"name": "ICT测试覆盖率", "value": "≥85%", "standard": "IPC-9252"},
                        {"name": "AOI缺陷检出率", "value": "≥95%", "standard": "IPC-A-610"},
                        {"name": "一次直通率 (FPY)", "value": "≥90%", "standard": "IPC-7912"},
                        {"name": "焊接合格率", "value": "≥98%", "standard": "IPC-A-610 Class 2"},
                        {"name": "外观缺陷率", "value": "≤1%", "standard": "IPC-A-610"},
                    ], ensure_ascii=False),
                ),
                TestStandard(
                    name="稳定性测试",
                    description="基于 JEDEC JESD22 / IPC-9701 标准的板卡长期运行可靠性评估，验证在温度、湿度应力下的稳定性和寿命。",
                    metrics=json.dumps([
                        {"name": "热循环温度范围", "value": "-40°C ~ +85°C", "standard": "JESD22-A104 / IPC-9701"},
                        {"name": "热循环次数", "value": "≥500 次", "standard": "IPC-9701 Class 2"},
                        {"name": "高温老化时间", "value": "48-72h (60-85°C满载)", "standard": "JESD22-A103"},
                        {"name": "湿热测试条件", "value": "85°C / 85%RH / 168h", "standard": "JESD22-A101"},
                        {"name": "MTBF (平均无故障时间)", "value": "≥50000 小时", "standard": "MIL-HDBK-217F"},
                        {"name": "温变速率", "value": "≤20°C/min", "standard": "IPC-9701A"},
                    ], ensure_ascii=False),
                ),
                TestStandard(
                    name="压力测试",
                    description="基于 HALT / IPC-9704 / SAE J1211 标准的板卡极限应力评估，确定工作极限和破坏极限。",
                    metrics=json.dumps([
                        {"name": "HALT温度范围", "value": "-100°C ~ +200°C", "standard": "IPC-9592 / HALT"},
                        {"name": "随机振动", "value": "2-10000Hz / 50Grms", "standard": "SAE J1211"},
                        {"name": "机械冲击", "value": "1500G / 0.5ms", "standard": "MIL-STD-202G"},
                        {"name": "PCB应变限值 (刚性板)", "value": "≤500 με", "standard": "IPC-9704 Type1"},
                        {"name": "温变速率 (HALT)", "value": "≥50°C/min", "standard": "IPC-9592"},
                        {"name": "电压应力范围", "value": "额定电压 ±20%", "standard": "ASTM F1596"},
                    ], ensure_ascii=False),
                ),
                TestStandard(
                    name="装壳测试",
                    description="基于 IPC-A-630 / IEC 60529 / IEC 60068 标准的板卡外壳装配质量与防护性能评估，涵盖外观、尺寸、防护等级、机械强度、电磁兼容和散热性能。",
                    metrics=json.dumps([
                        {"name": "外观检验合格率", "value": "≥98%", "standard": "IPC-A-630 Class 2"},
                        {"name": "尺寸公差符合度", "value": "≤±0.1mm (关键尺寸)", "standard": "IPC-A-630 / ISO 2768-m"},
                        {"name": "外壳防护等级", "value": "≥IP54", "standard": "IEC 60529 / GB/T 4208"},
                        {"name": "机械冲击耐受", "value": "30G / 11ms / 3轴6向", "standard": "IEC 60068-2-27"},
                        {"name": "随机振动耐受", "value": "5-500Hz / 2Grms / 每轴1h", "standard": "IEC 60068-2-64"},
                        {"name": "电磁屏蔽效能", "value": "≥40dB (30MHz-1GHz)", "standard": "IEEE 299 / MIL-STD-285"},
                    ], ensure_ascii=False),
                ),
            ]
            for s in defaults:
                db.add(s)
            db.commit()
            print(f"[INFO] 已创建 {len(defaults)} 条默认测试标准")
    except Exception as e:
        db.rollback()
        print(f"[WARN] 创建默认测试标准失败: {e}")
    finally:
        db.close()

    PORT = int(os.getenv("APP_PORT", os.getenv("PORT", "8100")))

    # ---- 启动前校验 ----
    # 1. 端口检测
    import socket as _sock
    _probe = _sock.socket(_sock.AF_INET, _sock.SOCK_STREAM)
    try:
        _probe.settimeout(1)
        _probe.bind(("0.0.0.0", PORT))
        _probe.close()
    except OSError as e:
        # 找出占用端口的进程
        _who = ""
        try:
            import subprocess
            _r = subprocess.run(["lsof", "-ti", f":{PORT}"], capture_output=True, text=True, timeout=3)
            if _r.stdout.strip():
                _pids = _r.stdout.strip().split("\n")
                for _p in _pids[:3]:
                    _ps = subprocess.run(["ps", "-p", _p, "-o", "command="], capture_output=True, text=True, timeout=3)
                    if _ps.stdout.strip():
                        _who += f"  PID {_p}: {_ps.stdout.strip()[:80]}\n"
        except Exception:
            pass
        print("=" * 60)
        print(f"  [ERROR] 端口 {PORT} 已被占用，无法启动")
        if _who:
            print(f"  占用进程:\n{_who.rstrip()}")
        print(f"  解决方案:")
        print(f"    1. 停止占用进程: lsof -ti:{PORT} | xargs kill")
        print(f"    2. 或设置其他端口: PORT=8101 python3 app.py")
        print("=" * 60)
        exit(1)

    # 2. 数据库连接预检
    print("  正在检查数据库连接...")
    _db_ok = False
    try:
        _test_db = SessionLocal()
        _test_db.execute(text("SELECT 1"))
        _db_ok = True
        _test_db.close()
    except Exception as e:
        _err_msg = str(e)
        if "Access denied" in _err_msg:
            print(f"  [ERROR] 数据库认证失败: 用户 '{DB_CONFIG['user']}' 密码错误")
        elif "Can't connect" in _err_msg or "Connection refused" in _err_msg:
            print(f"  [ERROR] 无法连接数据库 {DB_CONFIG['host']}:{DB_CONFIG['port']}")
            print(f"         请确认 MySQL 服务已启动")
        else:
            print(f"  [ERROR] 数据库连接异常: {_err_msg[:120]}")
        print(f"  数据库配置: {DB_CONFIG['host']}:{DB_CONFIG['port']}/{DB_CONFIG['database']}")
        print(f"  请检查 .env 文件中的 DB_HOST / DB_PORT / DB_USER / DB_PASSWORD")
        exit(1)

    if not _db_ok:
        exit(1)

    # ---- 优雅退出 ----
    import signal
    def _shutdown(signum, frame):
        print(f"\n[INFO] 收到信号 {signum}，正在关闭...")
        engine.dispose()
        print("[INFO] 数据库连接池已关闭")
        exit(0)
    signal.signal(signal.SIGTERM, _shutdown)
    signal.signal(signal.SIGINT, _shutdown)

    print("=" * 60)
    print("  硬件测试记录系统 - Hardware Test Record System")
    print(f"  后端 API 地址: http://127.0.0.1:{PORT}")
    print(f"  数据库: {DB_CONFIG['host']}:{DB_CONFIG['port']}/{DB_CONFIG['database']}")
    print("=" * 60)
    app.run(host="0.0.0.0", port=PORT, debug=False)
